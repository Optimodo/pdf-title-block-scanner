"""Optional client-portal document-list load and revision/title checks."""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from drawing_qa.docref import canonical_doc_ref
from drawing_qa.filename import parse_filename
from drawing_qa.models import CheckStatus, DocumentResult, record_issue
from drawing_qa.tokens import (
    is_allowed_first_revision,
    is_successor_revision,
    next_revision,
    parse_pc_revision,
    revision_rank,
)

LIST_SUFFIXES = {".xlsx", ".xlsm", ".csv"}
_HEADER_SCAN_ROWS = 12
_DEFAULT_FIRST = ("P01",)
_REPORT_STEM = re.compile(r"_\d{6}(?:-\d+)?$", re.I)


@dataclass(frozen=True)
class ConstructionUpgradeSpec:
    revision: str = "C01"


@dataclass
class DocumentListLayout:
    doc_ref_headers: list[str]
    title_headers: list[str]
    revision_headers: list[str]
    status_headers: list[str] = field(default_factory=list)
    skip_name_contains: list[str] = field(default_factory=list)
    prefer_name_contains: list[str] = field(default_factory=list)
    first_revisions: dict[str, list[str]] = field(default_factory=dict)
    search_keys: dict[str, list[str]] = field(default_factory=dict)
    status_map: dict[str, str] = field(default_factory=dict)
    project_status_maps: dict[str, dict[str, str]] = field(default_factory=dict)
    uploadable_statuses: list[str] = field(default_factory=list)
    construction_upgrade: dict[str, ConstructionUpgradeSpec] = field(default_factory=dict)
    enabled: bool = True


@dataclass
class PortalDocument:
    doc_ref: str
    revision: str
    title: str = ""
    status: str = ""


@dataclass
class DocumentListIndex:
    path: Path
    by_ref: dict[str, PortalDocument] = field(default_factory=dict)
    has_status: bool = False

    def get(self, doc_ref: str | None) -> PortalDocument | None:
        key = canonical_doc_ref(doc_ref)
        if not key:
            return None
        return self.by_ref.get(key)


def load_document_list_layout(raw: dict | None) -> DocumentListLayout:
    data = raw or {}
    headers = data.get("headers") or {}
    projects = data.get("projects") or {}
    first: dict[str, list[str]] = {}
    keys: dict[str, list[str]] = {}
    project_status: dict[str, dict[str, str]] = {}
    upgrades: dict[str, ConstructionUpgradeSpec] = {}
    default_upgrade = _construction_upgrade_spec(
        data.get("construction_upgrade"),
        ConstructionUpgradeSpec(),
        default_enabled=False,
    ) or ConstructionUpgradeSpec()
    for code, spec in projects.items():
        if not isinstance(spec, dict):
            continue
        project = str(code).strip().upper()
        first[project] = [
            str(item).strip()
            for item in (spec.get("first_revisions") or ["P01"])
            if str(item).strip()
        ] or list(_DEFAULT_FIRST)
        keys[project] = [
            str(item).strip()
            for item in (spec.get("search_keys") or [])
            if str(item).strip()
        ]
        extra = _status_lookup(spec.get("status_map"))
        if extra:
            project_status[project] = extra
        upgrade = _construction_upgrade_spec(
            spec.get("construction_upgrade"), default_upgrade, default_enabled=False
        )
        if upgrade is not None:
            upgrades[project] = upgrade
    return DocumentListLayout(
        enabled=bool(data.get("enabled", True)),
        doc_ref_headers=_string_list(headers.get("doc_ref")),
        title_headers=_string_list(headers.get("doc_title")),
        revision_headers=_string_list(headers.get("revision")),
        status_headers=_string_list(headers.get("status")),
        skip_name_contains=_string_list(data.get("skip_name_contains")),
        prefer_name_contains=_string_list(data.get("prefer_name_contains")),
        first_revisions=first,
        search_keys=keys,
        status_map=_status_lookup(data.get("status_map")),
        project_status_maps=project_status,
        uploadable_statuses=_string_list(data.get("uploadable_statuses"))
        or ["Status A", "Status B", "Status C", "A", "B", "C"],
        construction_upgrade=upgrades,
    )


def _string_list(value: object) -> list[str]:
    if not value:
        return []
    if isinstance(value, str):
        return [value]
    return [str(item).strip() for item in value if str(item).strip()]


def _status_key(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip().lower()


def _status_lookup(raw: object) -> dict[str, str]:
    if not isinstance(raw, dict):
        return {}
    mapped: dict[str, str] = {}
    for key, value in raw.items():
        if key is None or str(key).strip() in {"", "None", "none"}:
            continue
        mapped[_status_key(str(key))] = str(value or "").strip()
    return mapped


def _construction_upgrade_spec(
    raw: object,
    defaults: ConstructionUpgradeSpec,
    *,
    default_enabled: bool = True,
) -> ConstructionUpgradeSpec | None:
    """Parse construction_upgrade: true, false, or {enabled, revision}."""
    if raw is None:
        return defaults if default_enabled else None
    if isinstance(raw, bool):
        return defaults if raw else None
    if isinstance(raw, dict):
        if not bool(raw.get("enabled", True)):
            return None
        revision = str(raw.get("revision") or defaults.revision).strip() or defaults.revision
        return ConstructionUpgradeSpec(revision=revision)
    text = str(raw).strip().lower()
    if text in {"1", "true", "yes", "on"}:
        return defaults
    return None


def is_spreadsheet(path: Path) -> bool:
    return path.is_file() and path.suffix.lower() in LIST_SUFFIXES and not path.name.startswith("~$")


def _skip_candidate(path: Path, layout: DocumentListLayout) -> bool:
    name = path.name.lower()
    if any(token.lower() in name for token in layout.skip_name_contains):
        return True
    return bool(_REPORT_STEM.search(path.stem))


def _prefer_score(path: Path, layout: DocumentListLayout) -> int:
    name = path.name.lower()
    return sum(1 for token in layout.prefer_name_contains if token.lower() in name)


def find_document_list(
    folder: Path,
    layout: DocumentListLayout,
    *,
    explicit: Path | None = None,
    project_codes: list[str] | None = None,
) -> Path | None:
    """Return a portal dump path, or None when none looks usable."""
    if explicit is not None:
        return explicit if is_spreadsheet(explicit) else None
    if not layout.enabled or not folder.is_dir():
        return None
    candidates = [
        path
        for path in folder.iterdir()
        if is_spreadsheet(path) and not _skip_candidate(path, layout)
    ]
    if not candidates:
        return None
    scored: list[tuple[int, int, float, Path]] = []
    for path in candidates:
        mapping, _header_row, rows = _read_table(path, layout)
        if mapping is None or not rows:
            continue
        prefer = _prefer_score(path, layout)
        key_hits = 0
        stem = path.stem.upper()
        for code in project_codes or []:
            for key in layout.search_keys.get(code.upper(), []):
                if key.upper() in stem:
                    key_hits += 1
        scored.append((key_hits, prefer, path.stat().st_mtime, path))
    if not scored:
        return None
    scored.sort(key=lambda item: (item[0], item[1], item[2]), reverse=True)
    return scored[0][3]


def _norm_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def _header_index(headers: list[str], wanted: list[str]) -> int | None:
    lookup = {_norm_header(item): index for index, item in enumerate(headers)}
    for name in wanted:
        index = lookup.get(_norm_header(name))
        if index is not None:
            return index
    return None


def _map_headers(headers: list[str], layout: DocumentListLayout) -> dict[str, int] | None:
    doc_ref = _header_index(headers, layout.doc_ref_headers)
    revision = _header_index(headers, layout.revision_headers)
    if doc_ref is None or revision is None:
        return None
    mapping = {"doc_ref": doc_ref, "revision": revision}
    title = _header_index(headers, layout.title_headers)
    if title is not None:
        mapping["title"] = title
    if layout.status_headers:
        status = _header_index(headers, layout.status_headers)
        if status is not None:
            mapping["status"] = status
    return mapping


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "nan"}:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _doc_ref_from_cell(value: str) -> str:
    text = value.strip().strip(".")
    if not text:
        return ""
    parsed = parse_filename(text)
    if parsed.document_reference:
        return parsed.document_reference
    return text


def _read_csv_rows(path: Path) -> list[list[object]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return [list(row) for row in csv.reader(handle)]


def _read_excel_rows(path: Path) -> list[list[object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in wb.worksheets:
            rows = [
                [cell for cell in row]
                for row in sheet.iter_rows(max_row=5000, values_only=True)
            ]
            if rows:
                return rows
    finally:
        wb.close()
    return []


def _read_table(
    path: Path, layout: DocumentListLayout
) -> tuple[dict[str, int] | None, int, list[list[object]]]:
    try:
        rows = _read_csv_rows(path) if path.suffix.lower() == ".csv" else _read_excel_rows(path)
    except Exception:  # noqa: BLE001 - optional check must not abort the folder scan
        return None, 0, []
    for index, row in enumerate(rows[:_HEADER_SCAN_ROWS]):
        headers = [_cell_text(item) for item in row]
        mapping = _map_headers(headers, layout)
        if mapping:
            return mapping, index, rows
    return None, 0, []


def load_document_list(path: Path, layout: DocumentListLayout) -> DocumentListIndex:
    mapping, header_row, rows = _read_table(path, layout)
    index = DocumentListIndex(path=path, has_status="status" in (mapping or {}))
    if mapping is None:
        return index
    for row in rows[header_row + 1 :]:
        if not row:
            continue
        raw_ref = _cell_text(row[mapping["doc_ref"]] if mapping["doc_ref"] < len(row) else "")
        revision = _cell_text(row[mapping["revision"]] if mapping["revision"] < len(row) else "")
        title = ""
        if "title" in mapping and mapping["title"] < len(row):
            title = _cell_text(row[mapping["title"]])
        status = ""
        if "status" in mapping and mapping["status"] < len(row):
            status = _cell_text(row[mapping["status"]])
        doc_ref = canonical_doc_ref(_doc_ref_from_cell(raw_ref))
        if not doc_ref or not revision:
            continue
        incoming = PortalDocument(doc_ref=doc_ref, revision=revision, title=title, status=status)
        existing = index.by_ref.get(doc_ref)
        if existing is None or revision_rank(incoming.revision) > revision_rank(existing.revision):
            index.by_ref[doc_ref] = incoming
    return index


def first_revisions_for(project: str | None, layout: DocumentListLayout) -> list[str]:
    if project:
        listed = layout.first_revisions.get(project.strip().upper())
        if listed:
            return listed
    return list(_DEFAULT_FIRST)


def _local_doc_ref(result: DocumentResult) -> str | None:
    return result.titleblock.document_reference or result.filename.document_reference


def _local_revision(result: DocumentResult) -> str | None:
    return result.titleblock.revision or result.filename.revision


def _local_title(result: DocumentResult) -> str | None:
    return result.titleblock.title or result.filename.title


def intended_upload_revision(
    portal_rev: str | None,
    local_rev: str | None,
    *,
    require_revision: str | None = None,
) -> str:
    """Revision we will upload after designer corrections.

    If the drawing is already one issue up from the portal, keep that revision.
    If it is the same, skipped, or backwards (C01 on portal vs C03 on the drawing),
    use the next portal revision so document control is not asked to accept the wrong issue.
    When a project is upgrading P to C, require_revision (C01) wins.
    """
    if require_revision:
        return require_revision
    if portal_rev and local_rev and is_successor_revision(portal_rev, local_rev):
        return local_rev
    if portal_rev:
        return next_revision(portal_rev) or (local_rev or "")
    return local_rev or ""


def map_portal_status(raw: str, layout: DocumentListLayout, project: str | None) -> str:
    key = _status_key(raw)
    if not key:
        return ""
    if project:
        extra = layout.project_status_maps.get(project.strip().upper(), {})
        if key in extra:
            return extra[key]
    if key in layout.status_map:
        return layout.status_map[key]
    return raw.strip()


def status_letter(raw: str, layout: DocumentListLayout, project: str | None) -> str | None:
    """Return A, B, or C when the portal workflow status maps to one of those."""
    mapped = map_portal_status(raw, layout, project)
    for candidate in (mapped, raw):
        text = re.sub(r"\s+", " ", candidate or "").strip()
        if not text:
            continue
        match = re.fullmatch(r"(?:status\s*)?([ABC])", text, flags=re.IGNORECASE)
        if match:
            return match.group(1).upper()
    return None


def construction_upgrade_for(
    project: str | None, layout: DocumentListLayout
) -> ConstructionUpgradeSpec | None:
    if not project:
        return None
    return layout.construction_upgrade.get(project.strip().upper())


def should_upgrade_to_construction(
    portal: PortalDocument,
    layout: DocumentListLayout,
    project: str | None,
    *,
    has_status: bool,
) -> ConstructionUpgradeSpec | None:
    """Approved P issue on a project that has moved to construction issue."""
    spec = construction_upgrade_for(project, layout)
    if spec is None or not has_status:
        return None
    parsed = parse_pc_revision(portal.revision)
    if not parsed or parsed[0] != "P":
        return None
    if status_letter(portal.status, layout, project) not in {"A", "B"}:
        return None
    return spec


def _revision_matches(local: str | None, expected: str) -> bool:
    got = parse_pc_revision(local)
    want = parse_pc_revision(expected)
    return bool(got and want and got == want)


def status_allows_upload(raw: str, layout: DocumentListLayout, project: str | None) -> bool:
    """True when the current issue can be superseded (A/B/C or QA Approved)."""
    mapped = map_portal_status(raw, layout, project)
    for candidate in (mapped, raw):
        text = re.sub(r"\s+", " ", candidate or "").strip()
        if not text:
            continue
        if text.upper() in {"A", "B", "C"}:
            return True
        if re.fullmatch(r"status\s*[ABC]", text, flags=re.IGNORECASE):
            return True
        wanted = {_status_key(item) for item in layout.uploadable_statuses}
        if _status_key(text) in wanted:
            return True
    return False


def blocked_uploads(results: list[DocumentResult]) -> list[DocumentResult]:
    """PDFs that cannot be uploaded until document control sets portal status A/B/C.

    QA Approved is omitted: that workflow already allows a new revision.
    """
    return [item for item in results if item.portal_blocks_upload]


def _expected_first_note(allowed: list[str]) -> str:
    if len(allowed) == 1:
        return allowed[0]
    if len(allowed) == 2:
        return f"{allowed[0]} or {allowed[1]}"
    return ", ".join(allowed[:-1]) + f", or {allowed[-1]}"


def check_document_list(
    results: list[DocumentResult],
    index: DocumentListIndex,
    layout: DocumentListLayout,
) -> list[DocumentResult]:
    """Flag PDFs whose revision is not the next portal issue, or whose title disagrees."""
    source = index.path.name
    for result in results:
        result.portal_list_name = source
        doc_ref = _local_doc_ref(result)
        local_rev = _local_revision(result)
        project = (result.filename.parts.get("project") or "").strip().upper()
        allowed_first = first_revisions_for(project, layout)
        result.portal_first_revisions = allowed_first
        result.portal_has_status_column = index.has_status
        portal = index.get(doc_ref)
        if portal is None:
            if not local_rev:
                continue
            if is_allowed_first_revision(local_rev, allowed_first):
                continue
            expected = _expected_first_note(allowed_first)
            result.notes.append(
                f"Not on portal document list {source}; first issue should be {expected}, "
                f"not {local_rev}"
            )
            record_issue(result, CheckStatus.PORTAL_REVISION)
            continue

        result.portal_revision = portal.revision
        result.portal_title = portal.title or None
        result.portal_status = portal.status or None
        upgrade = should_upgrade_to_construction(
            portal, layout, project, has_status=index.has_status
        )
        if upgrade is not None:
            result.construction_upgrade_required = True
            result.proposed_upload_revision = intended_upload_revision(
                portal.revision, local_rev, require_revision=upgrade.revision
            )
        else:
            result.proposed_upload_revision = intended_upload_revision(
                portal.revision, local_rev
            )
        if index.has_status and not status_allows_upload(portal.status, layout, project):
            result.portal_blocks_upload = True
        if upgrade is not None:
            if not _revision_matches(local_rev, upgrade.revision):
                result.notes.append(
                    f"Portal list {source} has {doc_ref} at {portal.revision} "
                    f"({portal.status or 'no status'}); this project is issuing construction, "
                    f"so this file should be {upgrade.revision}, not {local_rev}"
                )
                record_issue(result, CheckStatus.PORTAL_REVISION)
        elif local_rev and not is_successor_revision(portal.revision, local_rev):
            nxt = next_revision(portal.revision)
            extra = ""
            if parse_pc_revision(portal.revision) and parse_pc_revision(portal.revision)[0] == "P":
                extra = " (or C01 if this is the first construction issue)"
            result.notes.append(
                f"Portal list {source} has {doc_ref} at {portal.revision}; "
                f"this file is {local_rev} (expected {nxt}{extra})"
            )
            record_issue(result, CheckStatus.PORTAL_REVISION)

        local_title = _local_title(result)
        if portal.title and local_title:
            from drawing_qa.compare import normalize_title

            portal_norm = normalize_title(portal.title)
            local_norm = normalize_title(local_title)
            if portal_norm and local_norm and portal_norm != local_norm:
                result.notes.append(
                    f"Portal list title {portal.title!r} does not match title-block "
                    f"title {local_title!r}"
                )
                record_issue(result, CheckStatus.PORTAL_TITLE)
    return results
