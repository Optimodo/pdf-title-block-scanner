from __future__ import annotations

from io import BytesIO
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from drawing_qa.designer_brief import (
    designer_actions,
    designer_doc_ref,
    designer_purpose_groups,
    designer_title,
)
from drawing_qa.document_list import blocked_uploads
from drawing_qa.dwg_pairing import find_dwg_files, unpaired_dwgs
from drawing_qa.models import CheckStatus, Confidence, DocumentResult
from drawing_qa.paths import (
    designer_report_path,
    document_control_report_path,
    next_available_paired_report_path,
    sanitize_filename_part,
)
from drawing_qa.preview import preview_size
from drawing_qa.timing import span as timing_span

STATUS_FILL = {
    CheckStatus.MATCH: PatternFill("solid", fgColor="C6EFCE"),
    CheckStatus.MISMATCH: PatternFill("solid", fgColor="FFC7CE"),
    CheckStatus.HISTORY_MISMATCH: PatternFill("solid", fgColor="F8CBAD"),
    CheckStatus.INCOMPLETE: PatternFill("solid", fgColor="FFEB9C"),
    CheckStatus.UNDETECTED: PatternFill("solid", fgColor="DDEBF7"),
    CheckStatus.SPELLING_ERROR: PatternFill("solid", fgColor="E4DFEC"),
    CheckStatus.DUPLICATE_REFERENCE: PatternFill("solid", fgColor="FFB6C1"),
    CheckStatus.DATE_REGRESSION: PatternFill("solid", fgColor="FFA07A"),
    CheckStatus.SUITABILITY_ERROR: PatternFill("solid", fgColor="F4B183"),
    CheckStatus.PURPOSE_MISMATCH: PatternFill("solid", fgColor="F8CBAD"),
    CheckStatus.PURPOSE_INCONSISTENT: PatternFill("solid", fgColor="F8CBAD"),
    CheckStatus.DWG_ISSUE: PatternFill("solid", fgColor="BDD7EE"),
    CheckStatus.PORTAL_REVISION: PatternFill("solid", fgColor="F4B183"),
    CheckStatus.PORTAL_TITLE: PatternFill("solid", fgColor="F8CBAD"),
    CheckStatus.CLIENT_ERROR: PatternFill("solid", fgColor="F8CBAD"),
    CheckStatus.FILENAME_PARSE_ERROR: PatternFill("solid", fgColor="F4B183"),
    CheckStatus.ERROR: PatternFill("solid", fgColor="D9D9D9"),
    CheckStatus.MULTIPLE_ISSUES: PatternFill("solid", fgColor="C65911"),
}
CONF_FILL = {
    Confidence.HIGH: PatternFill("solid", fgColor="C6EFCE"),
    Confidence.REVIEW: PatternFill("solid", fgColor="FFEB9C"),
}
BODY_FONT = Font(size=10)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
PURPOSE_KIND_FILL = PatternFill("solid", fgColor="DDEBF7")
PURPOSE_PROJECT_FILL = PatternFill("solid", fgColor="BDD7EE")
PURPOSE_VALUE_FILL = PatternFill("solid", fgColor="FFFFFF")
PURPOSE_KIND_FONT = Font(bold=True, size=10, color="1F4E79")
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")
LEFT_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
PREVIEW_W, PREVIEW_H = preview_size()
ROW_HEIGHT = 62
PREVIEW_COL_WIDTH = round(PREVIEW_W / 7.0, 1)

COLUMNS = [
    ("Status", 42),
    ("Confidence", 12),
    ("File (as scanned)", 35),
    ("Filename doc ref", 35),
    ("Title-block doc ref", 35),
    ("Filename title", 35),
    ("Title", 35),
    ("Rev (file)", 11),
    ("Rev (drawing)", 13),
    ("Status / suitability", 25),
    ("Client", 22),
    ("Date", 12),
    ("History latest", 40),
    ("History check", 14),
    ("Preview (detected fields)", PREVIEW_COL_WIDTH),
    ("New filename", 45),
    ("Rename result", 28),
    ("DWG pairing", 35),
    ("Notes", 60),
]
PREVIEW_COL = 15
HISTORY_CHECK_COL = 14
RENAME_RESULT_COL = 17
DWG_PAIRING_COL = 18

RENAME_FILL = {
    "Renamed": PatternFill("solid", fgColor="BDD7EE"),
    "Unchanged": PatternFill("solid", fgColor="C6EFCE"),
    "Not renamed": PatternFill("solid", fgColor="D9D9D9"),
    "Failed": PatternFill("solid", fgColor="FFC7CE"),
}
DWG_ISSUE_FILL = {
    "missing": PatternFill("solid", fgColor="FFC7CE"),
    "sheet_suffix": PatternFill("solid", fgColor="F8CBAD"),
    "name_differs": PatternFill("solid", fgColor="FFE699"),
    "orphan": PatternFill("solid", fgColor="DDEBF7"),
}
DWG_ISSUE_LABEL = {
    "missing": "Missing DWG",
    "sheet_suffix": "Sheet number .1 vs -1",
    "name_differs": "Paired, names differ",
    "orphan": "DWG with no PDF",
}


def _comp(result: DocumentResult, name: str):
    for item in result.comparisons:
        if item.name == name:
            return item
    return None


def _hcomp(result: DocumentResult, name: str):
    for item in result.history_comparisons:
        if item.name == name:
            return item
    return None


def _dwg_pairing_text(result: DocumentResult) -> str:
    """Generate text for DWG pairing column."""
    if result.paired_dwg:
        if result.dwg_issue == "sheet_suffix":
            return f"{result.paired_dwg.name} (.1 vs -1)"
        if result.dwg_mismatch:
            return f"{result.paired_dwg.name} (name mismatch)"
        return result.paired_dwg.name
    if result.dwg_files_present:
        return "No matching DWG"
    return "No DWGs in folder"


def _dwg_files_for(results: list[DocumentResult]) -> list[Path]:
    if not results or not any(item.dwg_files_present for item in results):
        return []
    return find_dwg_files(results[0].path.parent)


def _append_dwg_summary(ws: Worksheet, results: list[DocumentResult], row: int) -> int:
    """Write DWG pairing counts starting at row. Returns the next empty row."""
    dwgs = _dwg_files_for(results)
    exact = sum(1 for item in results if item.paired_dwg and not item.dwg_issue)
    suffix = sum(1 for item in results if item.dwg_issue == "sheet_suffix")
    names = sum(1 for item in results if item.dwg_issue == "name_differs")
    missing = sum(1 for item in results if item.dwg_issue == "missing")
    orphans = unpaired_dwgs(results, dwgs) if dwgs else []
    ws.cell(row, 1, "DWG pairing")
    ws.cell(row, 2, "Count")
    ws.cell(row, 3, "Meaning")
    for col in ("A", "B", "C"):
        ws[f"{col}{row}"].font = HEADER_FONT
        ws[f"{col}{row}"].fill = HEADER_FILL
    row += 1
    lines = [
        ("DWGs in folder", len(dwgs), "CAD files found next to the PDFs"),
        ("Paired, same name", exact, "PDF and DWG stems match"),
        (
            "Sheet number .1 vs -1",
            suffix,
            "Same drawing, but PDF uses 51333.1 and DWG uses 51333-1 (or the reverse)",
        ),
        ("Paired, names differ", names, "Same document reference; title or revision text differs"),
        ("PDFs missing a DWG", missing, "This folder has DWGs, but not one for this PDF"),
        ("DWGs with no PDF", len(orphans), "A DWG that did not pair with any scanned PDF"),
    ]
    if not dwgs:
        lines = [
            ("DWGs in folder", 0, "No DWG files here — pairing is skipped (often the whole set has none)"),
        ]
    for label, count, meaning in lines:
        ws.cell(row, 1, label)
        ws.cell(row, 2, count)
        ws.cell(row, 3, meaning)
        if label == "Sheet number .1 vs -1" and count:
            ws.cell(row, 1).fill = DWG_ISSUE_FILL["sheet_suffix"]
        elif label == "PDFs missing a DWG" and count:
            ws.cell(row, 1).fill = DWG_ISSUE_FILL["missing"]
        elif label == "DWGs with no PDF" and count:
            ws.cell(row, 1).fill = DWG_ISSUE_FILL["orphan"]
        ws.cell(row, 3).alignment = Alignment(wrap_text=True)
        for col in range(1, 4):
            ws.cell(row, col).font = BODY_FONT
        row += 1
    return row


def _write_dwg_sheet(ws: Worksheet, results: list[DocumentResult]) -> None:
    dwgs = _dwg_files_for(results)
    next_row = _append_dwg_summary(ws, results, 1)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 55
    ws.row_dimensions[1].height = 22

    row = next_row + 1
    headers = ("Issue", "PDF", "DWG", "Detail")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row, col, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    header_row = row
    row += 1

    issue_rows: list[tuple[str, str, str, str]] = []
    for result in results:
        issue = result.dwg_issue
        if not issue:
            continue
        issue_rows.append(
            (
                issue,
                result.path.name,
                result.paired_dwg.name if result.paired_dwg else "",
                "; ".join(
                    note
                    for note in result.notes
                    if "DWG" in note or "dwg" in note
                )
                or DWG_ISSUE_LABEL[issue],
            )
        )
    for dwg in unpaired_dwgs(results, dwgs):
        issue_rows.append(("orphan", "", dwg.name, "No PDF with a matching document reference"))

    if not dwgs:
        ws.cell(row, 1, "(none)")
        ws.cell(row, 3, "No DWG files in this folder, so missing CAD copies are not flagged.")
        ws.cell(row, 3).font = BODY_FONT
        return

    if not issue_rows:
        ws.cell(row, 1, "(none)")
        ws.cell(row, 3, "Every PDF that was scanned has a matching DWG with the same file name.")
        ws.cell(row, 3).font = BODY_FONT
        return

    for issue, pdf_name, dwg_name, detail in issue_rows:
        ws.cell(row, 1, DWG_ISSUE_LABEL[issue])
        ws.cell(row, 2, pdf_name)
        ws.cell(row, 3, dwg_name)
        ws.cell(row, 4, detail)
        fill = DWG_ISSUE_FILL.get(issue)
        if fill:
            ws.cell(row, 1).fill = fill
        for col in range(1, 5):
            ws.cell(row, col).font = BODY_FONT
            ws.cell(row, col).alignment = WRAP
        ws.row_dimensions[row].height = 28
        row += 1
    ws.auto_filter.ref = f"A{header_row}:D{row - 1}"
    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False


def _history_check(result: DocumentResult) -> str:
    if not result.titleblock.history.latest:
        return "No history"
    mismatches = [item for item in result.history_comparisons if item.matched is False]
    if mismatches:
        return "Mismatch"
    if any(item.detail.startswith("current field taken") for item in result.history_comparisons):
        return "From history"
    return "Matches current"


def _latest_history_label(result: DocumentResult) -> str:
    latest = result.titleblock.history.latest
    if not latest:
        return ""
    parts = [p for p in (latest.revision, latest.date, latest.suitability) if p]
    return " · ".join(parts)


def _rename_fill(result: DocumentResult) -> PatternFill | None:
    text = result.rename_result or ""
    if text.startswith("Renamed"):
        return RENAME_FILL["Renamed"]
    if text.startswith("Unchanged"):
        return RENAME_FILL["Unchanged"]
    if text.startswith("Failed"):
        return RENAME_FILL["Failed"]
    if text.startswith("Not renamed"):
        return RENAME_FILL["Not renamed"]
    return None


def _row(result: DocumentResult) -> list[object]:
    return [
        result.status_label(),
        result.confidence.value,
        result.original_filename or result.path.name,
        result.filename.document_reference or "",
        result.titleblock.document_reference or "",
        result.filename.title or "",
        result.titleblock.title or "",
        result.filename.revision or "",
        result.titleblock.revision or "",
        result.titleblock.suitability or "",
        result.titleblock.client or "",
        result.titleblock.date or "",
        _latest_history_label(result),
        _history_check(result),
        "",  # Preview column (filled separately)
        result.suggested_filename or "",
        result.rename_result or "",
        _dwg_pairing_text(result),
        "; ".join(result.notes + ([result.error] if result.error else [])),
    ]


def _style_header(ws: Worksheet) -> None:
    for col, (_name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    ws.sheet_view.showGridLines = False


def _add_preview(ws: Worksheet, row_idx: int, png: bytes, keep: list) -> None:
    buffer = BytesIO(png)
    buffer.name = "preview.png"
    image = XLImage(buffer)
    image.width = PREVIEW_W
    image.height = PREVIEW_H
    image.anchor = f"{get_column_letter(PREVIEW_COL)}{row_idx}"
    ws.add_image(image)
    keep.append(image)


def _write_rows(ws: Worksheet, results: list[DocumentResult], keep: list) -> None:
    ws.append([name for name, _width in COLUMNS])
    _style_header(ws)
    if not results:
        ws.append(["(none)"] + [""] * (len(COLUMNS) - 1))
        return
    for result in results:
        ws.append(_row(result))
        row_idx = ws.max_row
        ws.row_dimensions[row_idx].height = ROW_HEIGHT
        status_fill = STATUS_FILL.get(result.status)
        if result.status == CheckStatus.MULTIPLE_ISSUES:
            status_fill = STATUS_FILL[CheckStatus.MULTIPLE_ISSUES]
        if status_fill:
            ws.cell(row_idx, 1).fill = status_fill
        conf_fill = CONF_FILL.get(result.confidence)
        if conf_fill:
            ws.cell(row_idx, 2).fill = conf_fill
        history_fill = None
        check = _history_check(result)
        if check == "Mismatch":
            history_fill = STATUS_FILL[CheckStatus.HISTORY_MISMATCH]
        elif check == "Matches current":
            history_fill = STATUS_FILL[CheckStatus.MATCH]
        if history_fill:
            ws.cell(row_idx, HISTORY_CHECK_COL).fill = history_fill
        for col in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row_idx, col)
            cell.alignment = WRAP
            cell.font = BODY_FONT
        if result.preview_png:
            with timing_span("report_embed_previews"):
                _add_preview(ws, row_idx, result.preview_png, keep)
        rename_fill = _rename_fill(result)
        if rename_fill:
            ws.cell(row_idx, RENAME_RESULT_COL).fill = rename_fill
        if result.dwg_issue in DWG_ISSUE_FILL:
            ws.cell(row_idx, DWG_PAIRING_COL).fill = DWG_ISSUE_FILL[result.dwg_issue]
    last = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last}{ws.max_row}"


def _write_summary(ws: Worksheet, results: list[DocumentResult]) -> None:
    counts = Counter(item.status for item in results)
    conf_counts = Counter(item.confidence for item in results)
    ws["A1"] = "Drawing title-block QA"
    ws["A1"].font = Font(bold=True, size=10, color="1F4E79")
    ws["A2"] = "Generated"
    ws["B2"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws["A3"] = "Documents checked"
    ws["B3"] = len(results)
    ws["A4"] = "Start here"
    if any(item.confidence == Confidence.REVIEW for item in results):
        ws["B4"] = (
            "Send the designer workbook (or the Designer actions sheet) to CAD. "
            "Use Review needed for full detail and previews, then DWG pairing if this "
            "folder has CAD copies."
        )
    else:
        ws["B4"] = (
            "No designer actions. Use Review needed for full detail and previews, "
            "then DWG pairing if this folder has CAD copies."
        )
    ws["B4"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("B4:F4")
    for coord in ("A2", "B2", "A3", "B3", "A4", "B4"):
        ws[coord].font = BODY_FONT

    ws["A6"] = "Confidence"
    ws["B6"] = "Count"
    ws["A6"].font = HEADER_FONT
    ws["A6"].fill = HEADER_FILL
    ws["B6"].font = HEADER_FONT
    ws["B6"].fill = HEADER_FILL
    ws["A7"] = Confidence.REVIEW.value
    ws["B7"] = conf_counts.get(Confidence.REVIEW, 0)
    ws["A7"].fill = CONF_FILL[Confidence.REVIEW]
    ws["A7"].font = BODY_FONT
    ws["B7"].font = BODY_FONT
    ws["A8"] = Confidence.HIGH.value
    ws["B8"] = conf_counts.get(Confidence.HIGH, 0)
    ws["A8"].fill = CONF_FILL[Confidence.HIGH]
    ws["A8"].font = BODY_FONT
    ws["B8"].font = BODY_FONT

    ws["A10"] = "Status"
    ws["B10"] = "Count"
    ws["C10"] = "Meaning"
    for col in ("A", "B", "C"):
        ws[f"{col}10"].font = HEADER_FONT
        ws[f"{col}10"].fill = HEADER_FILL
    meanings = {
        CheckStatus.MATCH: "Filename and current title-block values agree; history revision/status match the latest row; date matches the first or latest history date",
        CheckStatus.MISMATCH: "Filename disagrees with the current title-block values (column A names the field)",
        CheckStatus.HISTORY_MISMATCH: "Current revision/status disagrees with the latest history row; the main date must match the first or latest history date",
        CheckStatus.INCOMPLETE: "Layout found, but the document reference could not be read from the title block",
        CheckStatus.UNDETECTED: "No configured layout scored high enough",
        CheckStatus.SPELLING_ERROR: "Possible spelling error detected in title",
        CheckStatus.DUPLICATE_REFERENCE: "Multiple PDFs have the same document reference",
        CheckStatus.DATE_REGRESSION: "Revision dates go backwards (later rev has earlier date)",
        CheckStatus.SUITABILITY_ERROR: "Purpose of issue / suitability is not in the whitelist",
        CheckStatus.PURPOSE_MISMATCH: "P revision with a construction purpose, or C revision with review and comment",
        CheckStatus.PURPOSE_INCONSISTENT: "No longer raised; off-list purposes are SUITABILITY_ERROR",
        CheckStatus.DWG_ISSUE: "DWG missing, or paired DWG uses -1 instead of .1 (or the reverse)",
        CheckStatus.PORTAL_REVISION: "Revision is not the next issue after the portal document list (or not a valid first issue if the drawing is new to the portal)",
        CheckStatus.PORTAL_TITLE: "Title disagrees with the portal document list",
        CheckStatus.CLIENT_ERROR: "Title-block client name is missing or not on the project list (clients.yaml)",
        CheckStatus.FILENAME_PARSE_ERROR: "Filename is not ISO 19650; title-block values are still shown",
        CheckStatus.ERROR: "PDF could not be read",
        CheckStatus.MULTIPLE_ISSUES: "More than one issue — see Notes and the status list in column A",
    }
    row = 11
    for status in CheckStatus:
        ws.cell(row, 1, status.value)
        ws.cell(row, 2, counts.get(status, 0))
        ws.cell(row, 3, meanings[status])
        fill = STATUS_FILL.get(status)
        if fill:
            ws.cell(row, 1).fill = fill
        ws.cell(row, 3).alignment = Alignment(wrap_text=True)
        for col in range(1, 4):
            ws.cell(row, col).font = BODY_FONT
        row += 1

    row = _append_dwg_summary(ws, results, row + 1)

    if any(item.rename_result for item in results):
        row += 1
        ws.cell(row, 1, "Rename")
        ws.cell(row, 2, "Count")
        ws.cell(row, 3, "Meaning")
        for col in ("A", "B", "C"):
            ws[f"{col}{row}"].font = HEADER_FONT
            ws[f"{col}{row}"].fill = HEADER_FILL
        row += 1
        renamed = sum(1 for item in results if (item.rename_result or "").startswith("Renamed"))
        unchanged = sum(1 for item in results if (item.rename_result or "").startswith("Unchanged"))
        skipped = sum(1 for item in results if (item.rename_result or "").startswith("Not renamed"))
        failed = sum(1 for item in results if (item.rename_result or "").startswith("Failed"))
        for label, count, fill in (
            ("Renamed", renamed, RENAME_FILL["Renamed"]),
            ("Unchanged", unchanged, RENAME_FILL["Unchanged"]),
            ("Not renamed / skipped", skipped, RENAME_FILL["Not renamed"]),
            ("Failed", failed, RENAME_FILL["Failed"]),
        ):
            ws.cell(row, 1, label)
            ws.cell(row, 2, count)
            ws.cell(row, 1).fill = fill
            for col in range(1, 3):
                ws.cell(row, col).font = BODY_FONT
            row += 1
        ws.cell(row, 1, "File (as scanned) is the name at the start of this run.")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1).font = BODY_FONT
        ws.cell(row, 1).alignment = Alignment(wrap_text=True)
        row += 1
        ws.cell(row, 1, "New filename is the name after this run (or the intended name if the rename was not applied).")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1).font = BODY_FONT
        ws.cell(row, 1).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 80
    ws.row_dimensions[4].height = 40


DESIGNER_COLUMNS = [
    ("Drawing number", 38),
    ("Title", 48),
    ("What to change", 95),
]
DESIGNER_HEADER_ROW = 8


def report_project_label(results: list[DocumentResult]) -> str:
    """Project name from whitelist settings, else the ISO project code."""
    labels: list[str] = []
    for result in results:
        code = (result.filename.parts.get("project") or "").strip()
        name = (result.purpose_list_name or "").strip()
        label = name or code
        if label and label not in labels:
            labels.append(label)
    return ", ".join(labels) if labels else "Drawings"


def report_stem(results: list[DocumentResult], *, when: datetime | None = None) -> str:
    """{Project}_{ddmmyy} used for both the main report and the designer sidecar."""
    when = when or datetime.now()
    label = sanitize_filename_part(report_project_label(results))
    return f"{label}_{when.strftime('%d%m%y')}"


def default_report_path(
    folder: Path,
    results: list[DocumentResult],
    *,
    when: datetime | None = None,
) -> Path:
    return next_available_paired_report_path(folder, report_stem(results, when=when))


def _write_simple_summary(ws: Worksheet, results: list[DocumentResult]) -> int:
    """Compact counts at the top of a designer sheet. Returns the actions header row."""
    label = report_project_label(results)
    when = datetime.now()
    total = len(results)
    need = sum(1 for item in results if item.confidence == Confidence.REVIEW)
    ok = sum(1 for item in results if item.confidence == Confidence.HIGH)
    title = ws.cell(1, 1, f"Designer actions — {label}")
    title.font = Font(bold=True, size=12, color="1F4E79")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    portal = next((item.portal_list_name for item in results if item.portal_list_name), "")
    rows = (
        (2, "Project", label),
        (3, "Date", when.strftime("%d/%m/%y")),
        (4, "Drawings checked", total),
        (5, "Need action", need),
        (6, "OK", ok),
        (7, "Portal list", portal or "—"),
    )
    for row, heading, value in rows:
        key = ws.cell(row, 1, heading)
        val = ws.cell(row, 2, value)
        key.font = BODY_FONT
        val.font = BODY_FONT
    ws.cell(5, 2).fill = CONF_FILL[Confidence.REVIEW]
    ws.cell(6, 2).fill = CONF_FILL[Confidence.HIGH]
    ws.row_dimensions[1].height = 22
    return DESIGNER_HEADER_ROW


def _wrapped_line_count(text: str, col_width: float) -> int:
    usable = max(int(col_width), 8)
    lines = 0
    for paragraph in str(text or "").split("\n"):
        if not paragraph:
            lines += 1
            continue
        current = 0
        para_lines = 1
        for word in paragraph.split(" "):
            add = len(word) + (1 if current else 0)
            if current and current + add > usable:
                para_lines += 1
                current = len(word)
            else:
                current += add
        lines += para_lines
    return max(lines, 1)


def _designer_row_height(values: list[object], widths: list[float]) -> float:
    lines = 1
    for value, width in zip(values, widths, strict=True):
        lines = max(lines, _wrapped_line_count(str(value or ""), width))
    return min(max(18, lines * 14.5 + 4), 180)


def _apply_designer_cell(cell, *, header: bool = False, changes: bool = False) -> None:
    cell.border = THIN_BORDER
    if header:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        return
    cell.font = BODY_FONT
    cell.alignment = LEFT_WRAP if changes else CENTER_WRAP


def _style_purpose_cell(cell, *, fill, font) -> None:
    cell.fill = fill
    cell.font = font
    cell.alignment = CENTER_WRAP
    cell.border = THIN_BORDER


def _write_purpose_kind_row(ws: Worksheet, row: int, official: bool) -> int:
    label = "Official list" if official else "Suggested list"
    cell = ws.cell(row, 1, label)
    _style_purpose_cell(cell, fill=PURPOSE_KIND_FILL, font=PURPOSE_KIND_FONT)
    ws.row_dimensions[row].height = 18
    return row + 1


def _write_purpose_list(ws: Worksheet, results: list[DocumentResult], start_row: int) -> None:
    groups = designer_purpose_groups(results)
    if not groups:
        return
    row = start_row
    title = ws.cell(row, 1, "Approved purposes of issue")
    title.fill = HEADER_FILL
    title.font = HEADER_FONT
    title.alignment = HEADER_ALIGN
    title.border = THIN_BORDER
    ws.row_dimensions[row].height = 20
    row += 1
    flags = {official for _label, _values, official in groups}
    uniform = len(flags) == 1
    if uniform:
        row = _write_purpose_kind_row(ws, row, next(iter(flags)))
    for label, values, official in groups:
        if not uniform:
            row = _write_purpose_kind_row(ws, row, official)
        if official:
            heading = ws.cell(row, 1, label)
            _style_purpose_cell(
                heading, fill=PURPOSE_PROJECT_FILL, font=Font(bold=True, size=10)
            )
            ws.row_dimensions[row].height = 18
            row += 1
        for value in values:
            cell = ws.cell(row, 1, value)
            _style_purpose_cell(cell, fill=PURPOSE_VALUE_FILL, font=BODY_FONT)
            ws.row_dimensions[row].height = 18
            row += 1


def _write_designer_sheet(
    ws: Worksheet,
    review: list[DocumentResult],
    all_results: list[DocumentResult],
) -> None:
    """Plain-language actions for designers — same rows as Review needed, no previews."""
    widths = [width for _name, width in DESIGNER_COLUMNS]
    for col, (_name, width) in enumerate(DESIGNER_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "548235"

    header_row = _write_simple_summary(ws, all_results)
    for col, (name, _width) in enumerate(DESIGNER_COLUMNS, start=1):
        _apply_designer_cell(ws.cell(header_row, col, name), header=True)
    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = f"A{header_row + 1}"

    if not review:
        empty_row = header_row + 1
        ws.cell(empty_row, 1, "(none)")
        ws.cell(empty_row, 3, "No drawings need designer action.")
        for col in range(1, 4):
            _apply_designer_cell(ws.cell(empty_row, col), changes=(col == 3))
        ws.row_dimensions[empty_row].height = 22
        _write_purpose_list(ws, all_results, empty_row + 2)
        return

    for result in review:
        actions = designer_actions(result)
        values = [designer_doc_ref(result), designer_title(result), actions]
        row_idx = ws.max_row + 1
        if row_idx <= header_row:
            row_idx = header_row + 1
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col, value)
            _apply_designer_cell(cell, changes=(col == 3))
            status_fill = STATUS_FILL.get(result.status)
            if result.status == CheckStatus.MULTIPLE_ISSUES:
                status_fill = STATUS_FILL[CheckStatus.MULTIPLE_ISSUES]
            if col == 1 and status_fill:
                cell.fill = status_fill
        ws.row_dimensions[row_idx].height = _designer_row_height(values, widths)

    last_data = ws.max_row
    last = get_column_letter(len(DESIGNER_COLUMNS))
    ws.auto_filter.ref = f"A{header_row}:{last}{last_data}"
    _write_purpose_list(ws, all_results, last_data + 2)


def write_designer_report(results: list[DocumentResult], output: Path) -> Path | None:
    """Single-tab workbook for email: simple summary + designer actions."""
    review = [item for item in results if item.confidence == Confidence.REVIEW]
    if not review:
        return None
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Designer actions"
    _write_designer_sheet(ws, review, results)
    wb.save(output)
    return output


DOCCONTROL_COLUMNS = [
    ("Drawing number", 38),
    ("Title", 48),
    ("Current revision", 16),
    ("Proposed revision", 18),
    ("Current portal status", 36),
    ("Please change to", 16),
]
DOCCONTROL_HEADER_ROW = 7
DOCCONTROL_PLEASE_CHANGE = "A, B, or C"


def _document_control_needed(results: list[DocumentResult]) -> list[DocumentResult]:
    """Drawings that document control must action. Empty means no sheet or sidecar."""
    if not any(item.portal_has_status_column for item in results):
        return []
    return blocked_uploads(results)


def write_document_control_report(
    results: list[DocumentResult], output: Path
) -> Path | None:
    """One-tab workbook for client document control: drawings that cannot be uploaded."""
    blocked = _document_control_needed(results)
    if not blocked:
        return None
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Document control"
    _write_document_control_sheet(ws, blocked, results)
    wb.save(output)
    return output


def _write_document_control_sheet(
    ws: Worksheet,
    blocked: list[DocumentResult],
    all_results: list[DocumentResult],
) -> None:
    last_col = len(DOCCONTROL_COLUMNS)
    widths = [width for _name, width in DOCCONTROL_COLUMNS]
    for col, (_name, width) in enumerate(DOCCONTROL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "C65911"

    label = report_project_label(all_results)
    when = datetime.now()
    title = ws.cell(1, 1, f"Document control — cannot upload — {label}")
    title.font = Font(bold=True, size=12, color="1F4E79")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    summary = (
        (2, "Project", label),
        (3, "Date", when.strftime("%d/%m/%y")),
        (4, "Need action", len(blocked)),
    )
    for row, heading, value in summary:
        key = ws.cell(row, 1, heading)
        val = ws.cell(row, 2, value)
        key.font = BODY_FONT
        val.font = BODY_FONT
    ws.cell(4, 2).fill = CONF_FILL[Confidence.REVIEW]
    note = ws.cell(
        5,
        1,
        "These drawings cannot be uploaded until the portal status is A, B, or C "
        "so the current issue can be superseded. Proposed revision is the next issue "
        "after the current portal revision, after any designer corrections.",
    )
    note.font = BODY_FONT
    note.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=last_col)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[5].height = 32

    header_row = DOCCONTROL_HEADER_ROW
    for col, (name, _width) in enumerate(DOCCONTROL_COLUMNS, start=1):
        _apply_designer_cell(ws.cell(header_row, col, name), header=True)
    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = f"A{header_row + 1}"

    if not blocked:
        empty_row = header_row + 1
        ws.cell(empty_row, 1, "(none)")
        ws.cell(empty_row, 5, "No drawings are blocked by portal status.")
        for col in range(1, last_col + 1):
            _apply_designer_cell(ws.cell(empty_row, col))
        ws.row_dimensions[empty_row].height = 22
        return

    for offset, result in enumerate(blocked):
        row_idx = header_row + 1 + offset
        values = [
            designer_doc_ref(result),
            designer_title(result),
            result.portal_revision or "",
            result.proposed_upload_revision
            or result.titleblock.revision
            or result.filename.revision
            or "",
            result.portal_status or "(blank)",
            DOCCONTROL_PLEASE_CHANGE,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col, value)
            _apply_designer_cell(cell)
            if col == 5:
                cell.fill = CONF_FILL[Confidence.REVIEW]
        ws.row_dimensions[row_idx].height = _designer_row_height(values, widths)


def write_report(results: list[DocumentResult], output: Path) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    keep: list = []
    wb._preview_images = keep  # prevent GC of BytesIO-backed images

    with timing_span("report_build"):
        summary = wb.active
        summary.title = "Summary"
        _write_summary(summary, results)

        review = [
            item for item in results if item.confidence == Confidence.REVIEW
        ]
        high = [item for item in results if item.confidence == Confidence.HIGH]

        if review:
            designer_sheet = wb.create_sheet("Designer actions")
            _write_designer_sheet(designer_sheet, review, results)
        blocked = _document_control_needed(results)
        if blocked:
            control_sheet = wb.create_sheet("Document control")
            _write_document_control_sheet(control_sheet, blocked, results)
        review_sheet = wb.create_sheet("Review needed")
        _write_rows(review_sheet, review, keep)
        dwg_sheet = wb.create_sheet("DWG pairing")
        _write_dwg_sheet(dwg_sheet, results)
        high_sheet = wb.create_sheet("High confidence")
        _write_rows(high_sheet, high, keep)
        all_sheet = wb.create_sheet("All documents")
        _write_rows(all_sheet, results, keep)

    with timing_span("report_save"):
        wb.save(output)
        write_designer_report(results, designer_report_path(output))
        write_document_control_report(results, document_control_report_path(output))
    return output
