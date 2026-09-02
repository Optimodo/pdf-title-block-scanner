from pathlib import Path

from drawing_qa.checker import check_pdf, check_paths, iter_pdfs
from drawing_qa.config_loader import load_config
from drawing_qa.models import CheckStatus, Confidence
from drawing_qa.report import DESIGNER_HEADER_ROW, write_report
from tests.pdf_fixtures import (
    write_bottom_right_pdf,
    write_bottom_strip_pdf,
    write_mbs_bottom_pdf,
    write_mbs_classic_pdf,
    write_mbs_right_pdf,
    write_mbs_right_portrait_pdf,
    write_plain_pdf,
)


def test_detects_bottom_right_and_matches(tmp_path: Path, config_dir: Path):
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P01",
    )
    result = check_pdf(pdf, load_config(config_dir))
    assert result.titleblock.layout_id == "bottom_right"
    assert result.titleblock.document_reference == "ABC-WXY-ZZ-00-DR-A-0001"
    assert result.titleblock.revision == "P01"
    assert result.titleblock.title == "Ground Floor GA"
    assert result.status == CheckStatus.MATCH


def test_filename_without_revision_or_title_can_still_match(tmp_path: Path, config_dir: Path):
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P01",
    )
    result = check_pdf(pdf, load_config(config_dir))
    assert result.filename.revision is None
    assert result.filename.title is None
    assert result.status == CheckStatus.MATCH
    assert result.confidence == Confidence.HIGH


def test_detects_bottom_strip_layout(tmp_path: Path, config_dir: Path):
    pdf = write_bottom_strip_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0102-C02.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0102",
        title="Roof Plan",
        revision="C02",
    )
    result = check_pdf(pdf, load_config(config_dir))
    assert result.titleblock.layout_id == "bottom_strip"
    assert result.status == CheckStatus.MATCH
    assert result.titleblock.title == "Roof Plan"


def test_reports_revision_mismatch(tmp_path: Path, config_dir: Path):
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P03",
    )
    result = check_pdf(pdf, load_config(config_dir))
    assert result.status == CheckStatus.MISMATCH
    revision = next(item for item in result.comparisons if item.name == "revision")
    assert revision.matched is False


def test_undetected_when_no_title_block(tmp_path: Path, config_dir: Path):
    pdf = write_plain_pdf(tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf")
    result = check_pdf(pdf, load_config(config_dir))
    assert result.status == CheckStatus.UNDETECTED


def test_filename_parse_error(tmp_path: Path, config_dir: Path):
    pdf = write_bottom_right_pdf(
        tmp_path / "A-101.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Plan",
        revision="P01",
    )
    result = check_pdf(pdf, load_config(config_dir))
    assert result.status == CheckStatus.FILENAME_PARSE_ERROR
    assert result.titleblock.layout_id == "bottom_right"
    assert result.titleblock.document_reference == "ABC-WXY-ZZ-00-DR-A-0001"
    assert result.titleblock.revision == "P01"
    assert result.titleblock.title == "Plan"
    assert result.filename.document_reference is None


def test_excel_report_created(tmp_path: Path, config_dir: Path):
    match_pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P01",
    )
    mismatch_pdf = write_bottom_strip_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0102-C01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0102",
        title="Roof Plan",
        revision="C02",
    )
    results = check_paths([match_pdf, mismatch_pdf], load_config(config_dir))
    output = write_report(results, tmp_path / "report.xlsx")
    assert output.is_file()
    from openpyxl import load_workbook

    wb = load_workbook(output)
    assert set(wb.sheetnames) == {
        "Summary",
        "Designer actions",
        "Review needed",
        "DWG pairing",
        "High confidence",
        "All documents",
    }
    assert wb.sheetnames[1] == "Designer actions"
    assert wb["All documents"].max_row == 3  # header + 2 rows
    assert wb["High confidence"].max_row == 2
    assert wb["Review needed"].max_row == 2
    designer = wb["Designer actions"]
    assert designer.max_row >= wb["Review needed"].max_row
    assert designer["A1"].value.startswith("Designer actions")
    assert designer["A2"].value == "Project"
    assert designer["A5"].value == "Need action"
    assert designer.cell(DESIGNER_HEADER_ROW, 1).value == "Drawing number"
    assert designer.cell(DESIGNER_HEADER_ROW, 2).value == "Title"
    assert designer.cell(DESIGNER_HEADER_ROW, 3).value == "What to change"
    assert not designer._images
    first = DESIGNER_HEADER_ROW + 1
    assert designer.cell(first, 1).value == "ABC-WXY-ZZ-00-DR-A-0102"
    assert designer.cell(first, 2).value == "Roof Plan"
    assert "revision" in (designer.cell(first, 3).value or "").lower()
    assert "C02" in (designer.cell(first, 3).value or "")
    assert "C01" in (designer.cell(first, 3).value or "")
    assert designer.cell(first, 1).alignment.vertical == "center"
    assert designer.cell(first, 3).border.left.style == "thin"
    assert designer.cell(first, 3).alignment.wrap_text is True
    purpose_cells = [
        str(cell.value)
        for row in designer.iter_rows(min_col=1, max_col=1, values_only=False)
        for cell in row
        if cell.value
    ]
    assert "Approved purposes of issue" in purpose_cells
    assert "Suggested list" in purpose_cells
    assert "Official list" not in purpose_cells
    assert "S5 - For Construction" in purpose_cells
    assert wb["All documents"]._images
    data = wb["All documents"]
    assert data["F1"].value == "Filename title"
    assert data["J1"].value == "Status / suitability"
    assert data["K1"].value == "Client"
    assert data.column_dimensions["C"].width == 35
    assert data.column_dimensions["D"].width == 35
    assert data.column_dimensions["E"].width == 35
    assert data.column_dimensions["F"].width == 35
    assert data.column_dimensions["G"].width == 35
    assert data.column_dimensions["J"].width == 25
    assert data.column_dimensions["M"].width == 40
    # Column P is New filename (45); Rename result is Q; Notes moved to S (60)
    assert data["P1"].value == "New filename"
    assert data["Q1"].value == "Rename result"
    assert data.column_dimensions["P"].width == 45
    assert data.column_dimensions["S"].width == 60
    assert data["A1"].font.size == 10
    assert data["C2"].font.size == 10
    assert data["A1"].alignment.horizontal == "center"
    assert data["P1"].alignment.horizontal == "center"


def test_history_latest_used_not_older_rows(tmp_path: Path, config_dir: Path):
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P03.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P03",
        date="15.06.24",
        suitability="S4",
        history=[
            ("P01", "12.01.24", "First issue"),
            ("P02", "03.03.24", "Updated for comment"),
            ("P03", "15.06.24", "S4 Construction"),
        ],
    )
    result = check_pdf(pdf, load_config(config_dir))
    assert result.titleblock.revision == "P03"
    assert result.titleblock.date == "15.06.24"
    assert result.titleblock.history.latest is not None
    assert result.titleblock.history.latest.revision == "P03"
    assert result.status == CheckStatus.MATCH
    assert result.confidence == Confidence.HIGH
    assert result.preview_png is None  # preview.all_files is false; MATCH rows skip crops


def test_history_accepts_original_issue_date_in_title_block(tmp_path: Path, config_dir: Path):
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P03.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P03",
        date="12.01.24",
        suitability="S4",
        history=[
            ("P01", "12.01.24", "First issue"),
            ("P02", "03.03.24", "Updated for comment"),
            ("P03", "15.06.24", "S4 Construction"),
        ],
    )
    result = check_pdf(pdf, load_config(config_dir))
    assert result.titleblock.revision == "P03"
    assert result.titleblock.date == "12.01.24"
    assert result.titleblock.history.latest is not None
    assert result.titleblock.history.latest.revision == "P03"
    assert result.status == CheckStatus.MATCH


def test_history_match_can_skip_preview_when_mismatch_only(tmp_path: Path, config_dir: Path):
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P03.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P03",
        date="15.06.24",
        suitability="S4",
        history=[
            ("P01", "12.01.24", "First issue"),
            ("P02", "03.03.24", "Updated for comment"),
            ("P03", "15.06.24", "S4 Construction"),
        ],
    )
    config = load_config(config_dir)
    assert config.preview is not None
    config.preview.all_files = False
    result = check_pdf(pdf, config)
    assert result.status == CheckStatus.MATCH
    assert result.preview_png is None


def test_history_mismatch_against_current_rev(tmp_path: Path, config_dir: Path):
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P03.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P03",
        date="15.06.24",
        suitability="S4",
        history=[
            ("P01", "12.01.24", "First issue"),
            ("P02", "03.03.24", "Updated for comment"),
        ],
    )
    result = check_pdf(pdf, load_config(config_dir))
    assert result.titleblock.revision == "P03"
    assert result.titleblock.history.latest is not None
    assert result.titleblock.history.latest.revision == "P02"
    assert result.status == CheckStatus.HISTORY_MISMATCH
    assert result.confidence == Confidence.REVIEW
    assert result.preview_png  # mismatch rows get a preview


def test_mixed_purpose_without_project_list_uses_default_whitelist(
    tmp_path: Path, config_dir: Path
):
    pdf1 = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-C01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Floor Plan",
        revision="C01",
        suitability="S4 - Construction",
    )
    pdf2 = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0002-C01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0002",
        title="Roof Plan",
        revision="C01",
        suitability="S5 - For Construction",
    )
    results = check_paths([pdf1, pdf2], load_config(config_dir))
    by_name = {item.path.name: item for item in results}
    off_list = by_name["ABC-WXY-ZZ-00-DR-A-0001-C01.pdf"]
    on_list = by_name["ABC-WXY-ZZ-00-DR-A-0002-C01.pdf"]
    assert CheckStatus.SUITABILITY_ERROR in off_list.issues
    assert off_list.preview_png
    assert CheckStatus.SUITABILITY_ERROR not in on_list.issues
    assert on_list.status == CheckStatus.MATCH
    assert on_list.preview_png is None
    assert all(CheckStatus.PURPOSE_INCONSISTENT not in item.issues for item in results)


def test_history_note_does_not_flag_purpose(tmp_path: Path, config_dir: Path):
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-C01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Floor Plan",
        revision="C01",
        date="15.06.24",
        suitability="S5 - For Construction",
        history=[
            ("P01", "12.01.24", "First issue"),
            ("C01", "15.06.24", "S3 - Bathroom first fix"),
        ],
    )
    result = check_pdf(pdf, load_config(config_dir))
    assert result.status == CheckStatus.MATCH
    assert CheckStatus.HISTORY_MISMATCH not in result.issues


def test_history_whitelist_purpose_must_match(tmp_path: Path, config_dir: Path):
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-C01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Floor Plan",
        revision="C01",
        date="15.06.24",
        suitability="S5 - For Construction",
        history=[
            ("P01", "12.01.24", "First issue"),
            ("C01", "15.06.24", "S4 - For Stage Approval"),
        ],
    )
    result = check_pdf(pdf, load_config(config_dir))
    assert result.status == CheckStatus.HISTORY_MISMATCH
    assert any("S4" in note for note in result.notes)


def test_detects_mbs_bottom_title_block(tmp_path: Path, config_dir: Path):
    pdf = write_mbs_bottom_pdf(
        tmp_path / "R459-MBS-CZ-XX-DR-W-55801-C01.pdf",
        document_reference="R459-MBS-CZ-XX-DR-W-55801",
        title="Dry Riser Schematic",
        revision="C01",
        suitability="S5 - Construction",
        date="27.08.26",
    )
    result = check_pdf(pdf, load_config(config_dir))
    assert result.titleblock.layout_id == "mbs_bottom"
    assert result.titleblock.document_reference == "R459-MBS-CZ-XX-DR-W-55801"
    assert result.titleblock.revision == "C01"
    assert result.titleblock.title == "Dry Riser Schematic"
    assert result.titleblock.suitability == "S5 - Construction"
    assert result.titleblock.date == "27.08.26"
    assert result.titleblock.client == "Berkeley Homes"
    assert result.titleblock.history.latest is not None
    assert result.titleblock.history.latest.revision == "C01"
    assert result.status == CheckStatus.MATCH


def test_detects_mbs_classic_title_block(tmp_path: Path, config_dir: Path):
    pdf = write_mbs_classic_pdf(
        tmp_path / "R459-MBS-DZ-XX-DR-W-54002-P02.pdf",
        document_reference="R459-MBS-DZ-XX-DR-W-54002",
        title="Oval Village - Block D1 - LTHW Schematic",
        revision="P02",
        suitability="S3 - Review and Comment",
        date="27/08/26",
    )
    result = check_pdf(pdf, load_config(config_dir))
    assert result.titleblock.layout_id == "mbs_classic"
    assert result.titleblock.document_reference == "R459-MBS-DZ-XX-DR-W-54002"
    assert result.titleblock.revision == "P02"
    assert result.titleblock.title == "Oval Village - Block D1 - LTHW Schematic"
    assert result.titleblock.suitability == "S3 - Review and Comment"
    assert result.titleblock.date == "27/08/26"
    assert result.titleblock.client == "Berkeley Homes"
    assert result.titleblock.history.latest is not None
    assert result.titleblock.history.latest.revision == "P02"
    assert result.status == CheckStatus.MATCH


def test_detects_mbs_right_title_block(tmp_path: Path, config_dir: Path):
    pdf = write_mbs_right_pdf(
        tmp_path / "R459-MBS-DZ-BA-DR-W-55100-P01.pdf",
        document_reference="R459-MBS-DZ-BA-DR-W-55100",
        title="Block D Commercial Sprinkler layout Level LG\nSheet 1 of 4",
        revision="P01",
        suitability="S3",
        date="14.08.26",
    )
    result = check_pdf(pdf, load_config(config_dir))
    assert result.titleblock.layout_id == "mbs_right"
    assert result.titleblock.document_reference == "R459-MBS-DZ-BA-DR-W-55100"
    assert result.titleblock.revision == "P01"
    assert result.titleblock.title == "Block D Commercial Sprinkler layout Level LG Sheet 1 of 4"
    assert result.titleblock.suitability == "S3 - REVIEW & COMMENT"
    assert result.titleblock.date == "14.08.26"
    assert result.titleblock.client == "Berkeley"
    assert result.titleblock.history.latest is not None
    assert result.titleblock.history.latest.revision == "P01"
    assert result.status == CheckStatus.MATCH


def test_detects_mbs_right_portrait_title_block(tmp_path: Path, config_dir: Path):
    pdf = write_mbs_right_portrait_pdf(
        tmp_path / "WCR-MBS-XX-ZZ-DR-E-6000-C04.pdf",
        document_reference="WCR-MBS-XX-ZZ-DR-E-6000",
        title="Typical Electrical Setting Out Elevations",
        revision="C04",
        suitability="S4 - Construction",
        date="11/05/2026",
        client="Seven Capital Woodrow",
    )
    result = check_pdf(pdf, load_config(config_dir))
    assert result.titleblock.layout_id == "mbs_right_portrait"
    assert result.titleblock.document_reference == "WCR-MBS-XX-ZZ-DR-E-6000"
    assert result.titleblock.revision == "C04"
    assert result.titleblock.title == "Typical Electrical Setting Out Elevations"
    assert result.titleblock.client == "Seven Capital Woodrow"
    assert CheckStatus.UNDETECTED not in result.issues


def test_real_mbs_drawing_if_present(config_dir: Path):
    sample = Path("test files") / "R459-MBS-DZ-BA-DR-W-55100.pdf"
    if not sample.is_file():
        return
    result = check_pdf(sample, load_config(config_dir))
    assert result.titleblock.layout_id == "mbs_right"
    assert result.titleblock.document_reference == "R459-MBS-DZ-BA-DR-W-55100"
    assert result.titleblock.revision == "P01"
    assert result.titleblock.date == "14.08.26"
    assert result.titleblock.suitability and result.titleblock.suitability.startswith("S3")
    assert result.titleblock.title and "Sprinkler" in result.titleblock.title
    assert result.titleblock.history.latest is not None
    assert result.titleblock.history.latest.revision == "P01"
    assert result.status == CheckStatus.MATCH
    assert result.confidence == Confidence.HIGH


def test_real_mbs_bottom_drawing_if_present(config_dir: Path):
    sample = Path("test files") / "R459-MBS-CZ-XX-DR-W-55801-C01.pdf"
    if not sample.is_file():
        return
    result = check_pdf(sample, load_config(config_dir))
    assert result.titleblock.layout_id == "mbs_bottom"
    assert result.titleblock.document_reference == "R459-MBS-CZ-XX-DR-W-55801"
    assert result.titleblock.revision == "C01"
    assert result.titleblock.title == "Dry Riser Schematic"
    assert result.titleblock.suitability
    assert "CONSTRUCTION" in result.titleblock.suitability.upper()


def test_real_mbs_classic_drawing_if_present(config_dir: Path):
    sample = Path("test files") / "R459-MBS-DZ-XX-DR-W-54002.pdf"
    if not sample.is_file():
        return
    result = check_pdf(sample, load_config(config_dir))
    assert result.titleblock.layout_id == "mbs_classic"
    assert result.titleblock.document_reference == "R459-MBS-DZ-XX-DR-W-54002"
    assert result.titleblock.revision == "P02"
    assert result.titleblock.title and "LTHW" in result.titleblock.title
    assert result.titleblock.suitability and result.titleblock.suitability.startswith("S3")
    assert result.status == CheckStatus.MATCH


def test_real_multiline_title_and_suitability_if_present(config_dir: Path):
    sample = Path("test files") / "J106309-MEP-02-ZZ-DR-X-600026.pdf"
    if not sample.is_file():
        return
    result = check_pdf(sample, load_config(config_dir))
    assert result.titleblock.title == "Electrical Layout - Apartment Type Z2 - L"
    assert result.titleblock.suitability == "S2 - Suitable For Tender"


def test_real_wcr_construction_status_and_date_if_present(config_dir: Path):
    sample = Path("test files") / (
        "WCR-MBS-B7-XX-DR-M-5301 - B7 - Mechanical Services Layout"
        " - Below Level 00 - Sheet 01 of 03_C01.pdf"
    )
    if not sample.is_file():
        return
    result = check_pdf(sample, load_config(config_dir))
    assert result.filename.title and result.filename.title.startswith("B7")
    assert result.titleblock.suitability == "A - CONSTRUCTION"
    assert result.titleblock.date == "01.06.2026"
    assert result.titleblock.history.latest is not None
    assert result.titleblock.history.latest.revision == "C01"
    assert "Current date taken from latest history row" not in result.notes


def test_real_hpa_history_rows_if_present(config_dir: Path):
    sample = Path("test files") / "HPA-MBS-D3-LG-DR-X-55103-C02.pdf"
    if not sample.is_file():
        return
    result = check_pdf(sample, load_config(config_dir))
    assert result.titleblock.history.latest is not None
    assert result.titleblock.history.latest.revision == "C02"
    assert len(result.titleblock.history.rows) >= 4


def test_iter_pdfs_is_not_recursive(tmp_path: Path):
    write_plain_pdf(tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf")
    nested = tmp_path / "nested"
    write_plain_pdf(nested / "ABC-WXY-ZZ-00-DR-A-0002-P01.pdf")
    found = iter_pdfs(tmp_path, recursive=False)
    assert [p.name for p in found] == ["ABC-WXY-ZZ-00-DR-A-0001-P01.pdf"]
    found_all = iter_pdfs(tmp_path, recursive=True)
    assert len(found_all) == 2


def test_real_wcr_b4_kitchen_overlapping_title_if_present(config_dir: Path):
    sample = Path(
        r"c:\Users\MikeMcLean\OneDrive - Malcolm Building Services Ltd"
        r"\Documents\MBS\Proj\WCR\up\02-09-26\kitchens"
        r"\WCR-MBS-B4-ZZ-DR-E-6201.pdf"
    )
    if not sample.is_file():
        return
    result = check_pdf(sample, load_config(config_dir))
    assert result.titleblock.layout_id == "mbs_right"
    assert result.titleblock.title.startswith("B4 - Kitchen Electrical Setting-out")
    assert "Layout" in (result.titleblock.title or "")
    assert "B4-1A" in (result.titleblock.title or "")

