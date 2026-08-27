from datetime import datetime
from pathlib import Path

from drawing_qa.models import CheckStatus, Confidence, DocumentResult, FilenameFields, TitleBlockFields
from drawing_qa.paths import designer_report_path, document_control_report_path
from drawing_qa.report import (
    DESIGNER_HEADER_ROW,
    DOCCONTROL_HEADER_ROW,
    report_project_label,
    report_stem,
    write_report,
)


def _result(*, project: str, name: str = "", confidence: Confidence = Confidence.HIGH) -> DocumentResult:
    return DocumentResult(
        path=Path(f"{project}-WXY-ZZ-00-DR-A-0001-P01.pdf"),
        filename=FilenameFields(
            raw_stem=f"{project}-WXY-ZZ-00-DR-A-0001-P01",
            document_reference=f"{project}-WXY-ZZ-00-DR-A-0001",
            revision="P01",
            parse_ok=True,
            parts={"project": project},
        ),
        titleblock=TitleBlockFields(
            document_reference=f"{project}-WXY-ZZ-00-DR-A-0001",
            title="Ground Floor GA",
            revision="P01",
        ),
        status=CheckStatus.MATCH if confidence == Confidence.HIGH else CheckStatus.MISMATCH,
        confidence=confidence,
        purpose_list_name=name,
    )


def test_report_label_prefers_whitelist_name():
    assert report_project_label([_result(project="R459", name="Oval C+D")]) == "Oval C+D"


def test_report_label_falls_back_to_project_code():
    assert report_project_label([_result(project="ABC")]) == "ABC"


def test_report_label_joins_distinct_projects():
    results = [
        _result(project="R459", name="Oval C+D"),
        _result(project="J106309", name="Barking Riverside"),
    ]
    assert report_project_label(results) == "Oval C+D, Barking Riverside"


def test_report_stem_uses_ddmmyy():
    when = datetime(2026, 8, 26, 9, 0)
    assert report_stem([_result(project="R459", name="Oval C+D")], when=when) == "Oval C+D_260826"
    assert report_stem([_result(project="ABC")], when=when) == "ABC_260826"


def test_write_report_also_writes_designer_sidecar(tmp_path: Path):
    result = _result(project="R459", name="Oval C+D", confidence=Confidence.REVIEW)
    result.issues = [CheckStatus.MISMATCH]
    output = write_report([result], tmp_path / "full.xlsx")
    side = designer_report_path(output)
    assert side.is_file()
    from openpyxl import load_workbook

    main = load_workbook(output)
    assert main.sheetnames[0] == "Summary"
    assert "Designer actions" in main.sheetnames
    designer = load_workbook(side)
    assert designer.sheetnames == ["Designer actions"]
    sheet = designer.active
    assert sheet["A1"].value == "Designer actions — Oval C+D"
    assert sheet["B4"].value == 1
    assert sheet["B5"].value == 1
    assert sheet["B6"].value == 0
    assert sheet["A7"].value == "Portal list"
    assert sheet["B7"].value == "—"
    assert sheet.cell(DESIGNER_HEADER_ROW, 1).value == "Drawing number"
    assert sheet.cell(DESIGNER_HEADER_ROW + 1, 1).value == "R459-WXY-ZZ-00-DR-A-0001"
    assert main["Designer actions"]["A1"].value == sheet["A1"].value
    assert not document_control_report_path(output).is_file()


def test_write_report_skips_designer_when_nothing_to_action(tmp_path: Path):
    result = _result(project="R459", name="Oval C+D", confidence=Confidence.HIGH)
    output = write_report([result], tmp_path / "full.xlsx")
    from openpyxl import load_workbook

    assert "Designer actions" not in load_workbook(output).sheetnames
    assert not designer_report_path(output).is_file()


def test_write_report_writes_document_control_sidecar(tmp_path: Path):
    result = _result(project="R459", name="Oval C+D", confidence=Confidence.HIGH)
    result.portal_list_name = "OVCD Document Listing.xlsx"
    result.portal_has_status_column = True
    result.portal_blocks_upload = True
    result.portal_revision = "P01"
    result.portal_status = "Pending QA Check"
    result.titleblock.revision = "P02"
    output = write_report([result], tmp_path / "full.xlsx")
    control = document_control_report_path(output)
    assert control.is_file()
    from openpyxl import load_workbook

    sheet = load_workbook(control).active
    assert sheet.title == "Document control"
    assert "Document control" in load_workbook(output).sheetnames
    assert sheet["A2"].value == "Project"
    assert sheet["B4"].value == 1
    assert sheet["A6"].value is None or sheet["A6"].value == ""
    assert sheet.cell(DOCCONTROL_HEADER_ROW, 1).value == "Drawing number"
    data = DOCCONTROL_HEADER_ROW + 1
    assert sheet.cell(data, 1).value == "R459-WXY-ZZ-00-DR-A-0001"
    assert sheet.cell(data, 2).value == "Ground Floor GA"
    assert sheet.cell(data, 3).value == "P01"
    assert sheet.cell(data, 4).value == "P02"
    assert sheet.cell(data, 5).value == "Pending QA Check"
    assert sheet.cell(data, 6).value == "A, B, or C"
    assert sheet.cell(DOCCONTROL_HEADER_ROW, 1).alignment.horizontal == "center"
    assert sheet.cell(data, 1).alignment.horizontal == "center"
    assert sheet.cell(data, 4).alignment.horizontal == "center"


def test_write_report_skips_document_control_when_nothing_to_action(tmp_path: Path):
    result = _result(project="R459", name="Oval C+D", confidence=Confidence.HIGH)
    result.portal_list_name = "OVCD Document Listing.xlsx"
    result.portal_has_status_column = True
    result.portal_blocks_upload = False
    result.portal_status = "QA Approved"
    output = write_report([result], tmp_path / "full.xlsx")
    from openpyxl import load_workbook

    assert "Document control" not in load_workbook(output).sheetnames
    assert not document_control_report_path(output).is_file()


def test_document_control_uses_intended_revision_not_wrong_drawing_rev(tmp_path: Path):
    result = _result(project="R459", name="Oval C+D", confidence=Confidence.REVIEW)
    result.portal_list_name = "OVCD Document Listing.xlsx"
    result.portal_has_status_column = True
    result.portal_blocks_upload = True
    result.portal_revision = "C01"
    result.portal_status = "Pending QA Check"
    result.proposed_upload_revision = "C02"
    result.titleblock.revision = "C03"
    result.filename.revision = "C03"
    output = write_report([result], tmp_path / "full.xlsx")
    from openpyxl import load_workbook

    sheet = load_workbook(document_control_report_path(output)).active
    data = DOCCONTROL_HEADER_ROW + 1
    assert sheet.cell(data, 3).value == "C01"
    assert sheet.cell(data, 4).value == "C02"
    main = load_workbook(output)["Document control"]
    assert main.cell(data, 4).value == "C02"


def test_designer_sheet_centers_all_but_changes_column(tmp_path: Path):
    result = _result(project="R459", name="Oval C+D", confidence=Confidence.REVIEW)
    result.issues = [CheckStatus.MISMATCH]
    output = write_report([result], tmp_path / "full.xlsx")
    from openpyxl import load_workbook

    sheet = load_workbook(designer_report_path(output)).active
    header_row = DESIGNER_HEADER_ROW
    data = header_row + 1
    for col in (1, 2, 3):
        assert sheet.cell(header_row, col).alignment.horizontal == "center"
    assert sheet.cell(data, 1).alignment.horizontal == "center"
    assert sheet.cell(data, 2).alignment.horizontal == "center"
    assert sheet.cell(data, 3).alignment.horizontal == "left"
