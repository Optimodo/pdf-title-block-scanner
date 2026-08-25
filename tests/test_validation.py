from pathlib import Path

from drawing_qa.checker import check_pdf, check_paths
from drawing_qa.config_loader import load_config
from drawing_qa.models import (
    CheckStatus,
    DocumentResult,
    FilenameFields,
    HistoryRow,
    RevisionHistory,
    TitleBlockFields,
)
from drawing_qa.validation import check_date_regression
from tests.pdf_fixtures import write_bottom_right_pdf


def test_detects_duplicate_document_references(tmp_path: Path, config_dir: Path):
    """Test that duplicate document references are detected."""
    # Create two PDFs with same document reference
    pdf1 = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Floor Plan",
        revision="P01",
    )
    pdf2 = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-C02.pdf",  # Different filename
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",  # Same doc ref
        title="Floor Plan",
        revision="C02",
    )
    
    config = load_config(config_dir)
    results = check_paths([pdf1, pdf2], config)
    
    # Both should be flagged as duplicates
    assert len(results) == 2
    assert all(r.status == CheckStatus.DUPLICATE_REFERENCE for r in results)
    assert all("Duplicate document reference" in " ".join(r.notes) for r in results)


def test_date_regression_in_history(tmp_path: Path, config_dir: Path):
    """Test that date mismatches between current and history are detected."""
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P03.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Floor Plan",
        revision="P03",
        date="10.01.24",
        history=[
            ("P01", "15.01.24", "First issue"),
            ("P02", "10.03.24", "Update"),
            ("P03", "05.05.24", "Final"),  # History date doesn't match current
        ],
    )
    
    config = load_config(config_dir)
    result = check_pdf(pdf, config)
    
    # History mismatch should be detected (current date doesn't match latest history)
    assert result.status == CheckStatus.HISTORY_MISMATCH
    # Just verify some notes exist about the mismatch
    assert len(result.notes) > 0


def test_filename_suggestion_for_mismatch(tmp_path: Path, config_dir: Path):
    """Test that filename suggestions are generated for mismatches."""
    # Use a valid ISO 19650 filename that doesn't match the title block
    pdf = write_bottom_right_pdf(
        tmp_path / "XYZ-DEF-AA-BB-DR-M-9999-P01.pdf",  # Valid 7-field format but wrong doc ref
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",  # Correct doc ref in title block
        title="Floor Plan",
        revision="P01",
    )
    
    config = load_config(config_dir)
    results = check_paths([pdf], config)
    result = results[0]
    
    assert result.status == CheckStatus.MISMATCH
    assert result.suggested_filename == "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf"


def test_dwg_pairing_exact_match(tmp_path: Path, config_dir: Path):
    """Test that DWG files with exact matching names are paired."""
    # Create PDF
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Floor Plan",
        revision="P01",
    )
    
    # Create matching DWG (just an empty file for testing)
    dwg = tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.dwg"
    dwg.write_text("")
    
    config = load_config(config_dir)
    results = check_paths([pdf], config)
    result = results[0]
    
    assert result.paired_dwg == dwg
    assert result.dwg_mismatch is False


def test_dwg_pairing_with_naming_mismatch(tmp_path: Path, config_dir: Path):
    """Test that DWG files with different separators are detected as mismatches."""
    # Create PDF with dashes
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Floor Plan",
        revision="P01",
    )
    
    # Create DWG with underscores instead of dashes
    dwg = tmp_path / "ABC_WXY_ZZ_00_DR_A_0001_P01.dwg"
    dwg.write_text("")
    
    config = load_config(config_dir)
    results = check_paths([pdf], config)
    result = results[0]
    
    assert result.paired_dwg == dwg
    assert result.dwg_mismatch is True
    assert any("paired by document reference" in note.lower() for note in result.notes)


def test_no_dwg_pairing_when_dwg_absent(tmp_path: Path, config_dir: Path):
    """Test that absence of DWG file is handled gracefully."""
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Floor Plan",
        revision="P01",
    )
    
    config = load_config(config_dir)
    results = check_paths([pdf], config)
    result = results[0]
    
    assert result.paired_dwg is None
    assert result.dwg_mismatch is False


def test_dwg_pairs_when_pdf_has_revision_and_dwg_does_not(tmp_path: Path, config_dir: Path):
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001_C08.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Floor Plan",
        revision="C08",
    )
    dwg = tmp_path / "ABC-WXY-ZZ-00-DR-A-0001.dwg"
    dwg.write_text("")
    results = check_paths([pdf], load_config(config_dir))
    assert results[0].paired_dwg == dwg
    assert results[0].dwg_mismatch is True


def test_dwg_pairs_when_dwg_has_title_and_revision_and_pdf_does_not(
    tmp_path: Path, config_dir: Path
):
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Floor Plan",
        revision="P01",
    )
    dwg = tmp_path / "ABC-WXY-ZZ-00-DR-A-0001 - Floor Plan_P01.dwg"
    dwg.write_text("")
    results = check_paths([pdf], load_config(config_dir))
    assert results[0].paired_dwg == dwg
    assert results[0].dwg_mismatch is True


def test_dotted_sheet_suffixes_are_not_duplicates(tmp_path: Path, config_dir: Path):
    pdf1 = write_bottom_right_pdf(
        tmp_path / "R459-MBS-DZ-ZZ-DR-W-51333.1 - Sheet One.pdf",
        document_reference="R459-MBS-DZ-ZZ-DR-W-51333.1",
        title="Sheet One",
        revision="C01",
    )
    pdf2 = write_bottom_right_pdf(
        tmp_path / "R459-MBS-DZ-ZZ-DR-W-51333.2 - Sheet Two.pdf",
        document_reference="R459-MBS-DZ-ZZ-DR-W-51333.2",
        title="Sheet Two",
        revision="C01",
    )
    results = check_paths([pdf1, pdf2], load_config(config_dir))
    assert all(r.status == CheckStatus.MATCH for r in results)
    assert results[0].titleblock.document_reference == "R459-MBS-DZ-ZZ-DR-W-51333.1"
    assert results[1].titleblock.document_reference == "R459-MBS-DZ-ZZ-DR-W-51333.2"
    assert results[0].filename.document_reference == "R459-MBS-DZ-ZZ-DR-W-51333.1"
    assert results[1].filename.document_reference == "R459-MBS-DZ-ZZ-DR-W-51333.2"


def test_dwg_pairs_dotted_pdf_number_with_hyphen_dwg(tmp_path: Path, config_dir: Path):
    pdf = write_bottom_right_pdf(
        tmp_path / "R459-MBS-DZ-ZZ-DR-W-51333.1 - Sheet One.pdf",
        document_reference="R459-MBS-DZ-ZZ-DR-W-51333.1",
        title="Sheet One",
        revision="C01",
    )
    dwg = tmp_path / "R459-MBS-DZ-ZZ-DR-W-51333-1.dwg"
    dwg.write_text("")
    results = check_paths([pdf], load_config(config_dir))
    assert results[0].paired_dwg == dwg
    assert results[0].dwg_mismatch is True
    assert results[0].dwg_issue == "sheet_suffix"
    assert CheckStatus.DWG_ISSUE in results[0].issues
    assert any(".1 vs -1" in note for note in results[0].notes)


def test_missing_dwg_is_flagged_when_folder_has_other_dwgs(tmp_path: Path, config_dir: Path):
    pdf_ok = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Floor Plan",
        revision="P01",
    )
    pdf_missing = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0002-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0002",
        title="Roof Plan",
        revision="P01",
    )
    (tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.dwg").write_text("")
    results = check_paths([pdf_ok, pdf_missing], load_config(config_dir))
    by_name = {item.path.name: item for item in results}
    assert by_name[pdf_ok.name].paired_dwg is not None
    assert by_name[pdf_ok.name].dwg_issue is None
    assert CheckStatus.DWG_ISSUE not in by_name[pdf_ok.name].issues
    assert by_name[pdf_missing.name].paired_dwg is None
    assert by_name[pdf_missing.name].dwg_issue == "missing"
    assert CheckStatus.DWG_ISSUE in by_name[pdf_missing.name].issues


def test_dwg_report_tab_lists_sheet_suffix_and_missing(tmp_path: Path, config_dir: Path):
    from openpyxl import load_workbook

    from drawing_qa.report import write_report

    pdf_suffix = write_bottom_right_pdf(
        tmp_path / "R459-MBS-DZ-ZZ-DR-W-51333.1 - Sheet One.pdf",
        document_reference="R459-MBS-DZ-ZZ-DR-W-51333.1",
        title="Sheet One",
        revision="C01",
    )
    pdf_missing = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0002-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0002",
        title="Roof Plan",
        revision="P01",
    )
    (tmp_path / "R459-MBS-DZ-ZZ-DR-W-51333-1.dwg").write_text("")
    (tmp_path / "ORPHAN-DWG.dwg").write_text("")
    results = check_paths([pdf_suffix, pdf_missing], load_config(config_dir))
    output = write_report(results, tmp_path / "report.xlsx")
    wb = load_workbook(output)
    assert "DWG pairing" in wb.sheetnames
    sheet = wb["DWG pairing"]
    texts = [
        str(cell.value)
        for row in sheet.iter_rows(min_row=1, max_col=4, values_only=False)
        for cell in row
        if cell.value
    ]
    blob = " ".join(texts)
    assert "Sheet number .1 vs -1" in blob
    assert "Missing DWG" in blob
    assert "DWG with no PDF" in blob
    assert "ORPHAN-DWG.dwg" in blob


def test_duplicates_dont_override_serious_issues(tmp_path: Path, config_dir: Path):
    """Test that duplicate detection doesn't override more serious statuses."""
    # Create two PDFs with same doc ref but one has a mismatch
    pdf1 = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Floor Plan",
        revision="P01",
    )
    pdf2 = write_bottom_right_pdf(
        tmp_path / "XYZ-DEF-AA-BB-DR-M-9999-C02.pdf",  # Valid 7-field format but wrong doc ref
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",  # Same doc ref (duplicate)
        title="Floor Plan",
        revision="C02",
    )
    
    config = load_config(config_dir)
    results = check_paths([pdf1, pdf2], config)
    
    # pdf1: duplicate only (filename matches its title block)
    # pdf2: mismatch plus duplicate
    assert results[0].status == CheckStatus.DUPLICATE_REFERENCE
    assert results[1].status == CheckStatus.MULTIPLE_ISSUES
    assert CheckStatus.MISMATCH in results[1].issues
    assert CheckStatus.DUPLICATE_REFERENCE in results[1].issues
    
    assert all("Duplicate document reference" in " ".join(r.notes) for r in results)


def _history_result(rows: list[HistoryRow], *, current_date: str | None = None) -> DocumentResult:
    latest = max(rows, key=lambda row: row.revision or "")
    return DocumentResult(
        path=Path("sheet.pdf"),
        filename=FilenameFields(raw_stem="sheet", parse_ok=True),
        titleblock=TitleBlockFields(
            document_reference="ABC-WXY-ZZ-00-DR-A-0001",
            date=current_date,
            history=RevisionHistory(rows=rows, latest=latest),
        ),
        status=CheckStatus.MATCH,
    )


def test_date_regression_uses_revision_rank_not_visual_order():
    """Newest row at the top (upward-growing table) is not a regression."""
    from drawing_qa.models import finalize_status

    result = _history_result(
        [
            HistoryRow(revision="C03", date="11.05.2026"),
            HistoryRow(revision="C02", date="01.04.2026"),
            HistoryRow(revision="P01", date="15.01.2026"),
        ],
        current_date="11.05.2026",
    )
    check_date_regression([result])
    finalize_status(result)
    assert result.status == CheckStatus.MATCH
    assert CheckStatus.DATE_REGRESSION not in result.issues
    assert not any("Date regression" in note for note in result.notes)


def test_date_regression_still_detected_when_later_rev_is_older_date():
    from drawing_qa.models import finalize_status

    result = _history_result(
        [
            HistoryRow(revision="P02", date="01.01.24"),
            HistoryRow(revision="P01", date="15.06.24"),
        ]
    )
    check_date_regression([result])
    finalize_status(result)
    assert CheckStatus.DATE_REGRESSION in result.issues
    assert any("Date regression" in note for note in result.notes)


def test_date_regression_does_not_compare_p_series_with_c_series():
    from drawing_qa.models import finalize_status

    result = _history_result(
        [
            HistoryRow(revision="P03", date="21.08.26"),
            HistoryRow(revision="C02", date="11.08.26"),
            HistoryRow(revision="C01", date="16.06.26"),
            HistoryRow(revision="P02", date="29.05.26"),
        ],
        current_date="21.08.26",
    )
    check_date_regression([result])
    finalize_status(result)
    assert CheckStatus.DATE_REGRESSION not in result.issues
