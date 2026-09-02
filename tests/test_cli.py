from datetime import datetime
from pathlib import Path

from drawing_qa.cli import main
from tests.pdf_fixtures import write_bottom_right_pdf


def _abc_report_stem() -> str:
    return f"ABC_{datetime.now().strftime('%d%m%y')}"


def test_cli_check_writes_report(tmp_path: Path, config_dir: Path):
    drawings = tmp_path / "drawings"
    write_bottom_right_pdf(
        drawings / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P01",
    )
    output = tmp_path / "out.xlsx"
    code = main(
        [
            "check",
            str(drawings),
            "--config-dir",
            str(config_dir),
            "--output",
            str(output),
            "--no-pause",
        ]
    )
    assert code == 0
    assert output.is_file()
    assert not (tmp_path / "out_designer.xlsx").is_file()


def test_cli_inspect_writes_crops(tmp_path: Path, config_dir: Path):
    pdf = write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P01",
    )
    debug = tmp_path / "debug"
    code = main(
        [
            "inspect",
            str(pdf),
            "--config-dir",
            str(config_dir),
            "--debug-dir",
            str(debug),
            "--no-pause",
        ]
    )
    assert code == 0
    assert (debug / "ABC-WXY-ZZ-00-DR-A-0001-P01_bottom_right.png").is_file()
    assert (debug / "ABC-WXY-ZZ-00-DR-A-0001-P01_bottom_strip.png").is_file()


def test_auto_check_uses_program_folder(tmp_path: Path, monkeypatch):
    write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P01",
    )
    monkeypatch.setattr("drawing_qa.cli.app_dir", lambda: tmp_path)
    code = main(["--no-pause"])
    assert code == 0
    stem = _abc_report_stem()
    assert (tmp_path / f"{stem}.xlsx").is_file()
    assert not (tmp_path / f"{stem}_designer.xlsx").is_file()


def test_auto_check_versions_existing_report(tmp_path: Path, monkeypatch):
    write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P01",
    )
    stem = _abc_report_stem()
    (tmp_path / f"{stem}.xlsx").write_bytes(b"old")
    monkeypatch.setattr("drawing_qa.cli.app_dir", lambda: tmp_path)
    code = main(["--no-pause"])
    assert code == 0
    assert (tmp_path / f"{stem}-1.xlsx").is_file()
    assert not (tmp_path / f"{stem}-1_designer.xlsx").is_file()
    assert (tmp_path / f"{stem}.xlsx").read_bytes() == b"old"


def test_dropped_excel_is_used_as_document_list(tmp_path: Path, monkeypatch, capsys):
    from openpyxl import Workbook

    write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P02.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P02",
    )
    listing = tmp_path / "OVCD Document Listing.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Original Doc Ref (Non-Standard)", "Description", "Revision"])
    ws.append(["ABC-WXY-ZZ-00-DR-A-0001", "Ground Floor GA", "P01"])
    wb.save(listing)
    monkeypatch.setattr("drawing_qa.cli.app_dir", lambda: tmp_path)
    code = main([str(listing), "--no-pause"])
    assert code == 0
    stem = _abc_report_stem()
    from openpyxl import load_workbook

    captured = capsys.readouterr()
    assert "OVCD Document Listing.xlsx" in captured.out
    assert load_workbook(tmp_path / f"{stem}.xlsx")["All documents"]["A2"].value == "MATCH"
    assert not (tmp_path / f"{stem}_designer.xlsx").is_file()


def test_check_spreadsheet_argument_scans_parent_folder(tmp_path: Path, config_dir: Path):
    from openpyxl import Workbook

    drawings = tmp_path / "drawings"
    write_bottom_right_pdf(
        drawings / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P01",
    )
    listing = drawings / "Document Listing.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Original Doc Ref (Non-Standard)", "Description", "Revision"])
    ws.append(["ABC-WXY-ZZ-00-DR-A-0001", "Ground Floor GA", "P01"])
    wb.save(listing)
    output = tmp_path / "out.xlsx"
    code = main(
        [
            "check",
            str(listing),
            "--config-dir",
            str(config_dir),
            "--output",
            str(output),
            "--no-pause",
        ]
    )
    assert code == 1
    from openpyxl import load_workbook

    status = load_workbook(output)["All documents"]["A2"].value
    assert "PORTAL_REVISION" in str(status)


def test_list_checks_prints_ids(capsys):
    code = main(["--list-checks", "--no-pause"])
    assert code == 0
    out = capsys.readouterr().out
    assert "portal-revision" in out
    assert "mismatch" in out


def test_check_disable_portal_revision(tmp_path: Path, config_dir: Path):
    from openpyxl import Workbook, load_workbook

    drawings = tmp_path / "drawings"
    write_bottom_right_pdf(
        drawings / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P01",
    )
    listing = drawings / "Document Listing.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Original Doc Ref (Non-Standard)", "Description", "Revision"])
    ws.append(["ABC-WXY-ZZ-00-DR-A-0001", "Ground Floor GA", "P01"])
    wb.save(listing)
    output = tmp_path / "out.xlsx"
    code = main(
        [
            "check",
            str(listing),
            "--config-dir",
            str(config_dir),
            "--output",
            str(output),
            "--disable",
            "portal-revision",
            "--no-pause",
        ]
    )
    assert code == 0
    status = load_workbook(output)["All documents"]["A2"].value
    assert status == "MATCH"


def test_unknown_check_exits_2(capsys):
    code = main(["--disable", "not-a-check", "--no-pause"])
    assert code == 2
    err = capsys.readouterr().err
    assert "Unknown check" in err


def test_custom_entry_prints_banner(tmp_path: Path, monkeypatch, capsys):
    from drawing_qa.cli_custom import main as custom_main

    write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P01",
    )
    monkeypatch.setattr("drawing_qa.cli.app_dir", lambda: tmp_path)
    code = custom_main(["--disable", "portal", "--no-pause"])
    assert code == 0
    out = capsys.readouterr().out
    assert "TBCheckCustom" in out
    assert "portal-revision" in out


def test_custom_prompt_toggles_portal_revision(tmp_path: Path, monkeypatch, capsys):
    from drawing_qa.cli_custom import main as custom_main

    write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P01",
    )
    monkeypatch.setattr("drawing_qa.cli.app_dir", lambda: tmp_path)
    answers = iter(["9", ""])
    monkeypatch.setattr("builtins.input", lambda *a, **k: next(answers))
    code = custom_main(["--prompt-checks", "--no-pause"])
    assert code == 0
    out = capsys.readouterr().out
    assert "12 QA checks" in out
    assert "Disabled checks: portal-revision" in out



