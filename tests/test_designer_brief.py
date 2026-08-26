from pathlib import Path

from drawing_qa.compare import build_result
from drawing_qa.designer_brief import designer_actions, designer_doc_ref, designer_title
from drawing_qa.models import (
    CheckStatus,
    Confidence,
    DocumentResult,
    FieldComparison,
    FilenameFields,
    HistoryRow,
    RevisionHistory,
    TitleBlockFields,
)

COMPARE_RULES = {
    "document_reference": "required",
    "revision": "if_both_present",
    "title": "if_both_present",
    "suitability": "if_both_present",
    "date": "if_both_present",
}


def _base(
    *,
    filename_rev: str = "P01",
    titleblock_rev: str = "P01",
    title: str = "Ground Floor GA",
    suitability: str | None = None,
) -> DocumentResult:
    return DocumentResult(
        path=Path("ABC-WXY-ZZ-00-DR-A-0001-P01.pdf"),
        filename=FilenameFields(
            raw_stem="ABC-WXY-ZZ-00-DR-A-0001-P01",
            document_reference="ABC-WXY-ZZ-00-DR-A-0001",
            revision=filename_rev,
            parse_ok=True,
        ),
        titleblock=TitleBlockFields(
            layout_id="bottom_right",
            document_reference="ABC-WXY-ZZ-00-DR-A-0001",
            revision=titleblock_rev,
            title=title,
            suitability=suitability,
        ),
    )


def test_match_needs_no_action():
    result = build_result(_base(), COMPARE_RULES)
    assert result.status == CheckStatus.MATCH
    assert designer_actions(result) == "No designer action required."


def test_revision_mismatch_tells_designer_to_change_title_block():
    result = build_result(_base(filename_rev="P01", titleblock_rev="P03"), COMPARE_RULES)
    text = designer_actions(result)
    assert "revision" in text.lower()
    assert "'P03'" in text
    assert "'P01'" in text
    assert "title block" in text.lower()
    assert designer_doc_ref(result) == "ABC-WXY-ZZ-00-DR-A-0001"
    assert designer_title(result) == "Ground Floor GA"


def test_doc_ref_mismatch_states_filename_and_both_numbers():
    result = build_result(
        DocumentResult(
            path=Path("R459-MBS-DZ-ZZ-DR-W-55056.pdf"),
            filename=FilenameFields(
                raw_stem="R459-MBS-DZ-ZZ-DR-W-55056",
                document_reference="R459-MBS-DZ-ZZ-DR-W-55056",
                parse_ok=True,
            ),
            titleblock=TitleBlockFields(
                layout_id="bottom_right",
                document_reference="R459-MBS-DZ-ZZ-DR-W-55053",
                title="Apartment Type D1.8.2",
                revision="P04",
            ),
            original_filename="R459-MBS-DZ-ZZ-DR-W-55056.pdf",
        ),
        COMPARE_RULES,
    )
    text = designer_actions(result)
    assert "R459-MBS-DZ-ZZ-DR-W-55053" in text
    assert "R459-MBS-DZ-ZZ-DR-W-55056" in text
    assert "should match" in text.lower()
    assert "one of them needs changing" in text.lower()
    assert "change the title-block drawing number to" not in text.lower()


def test_history_date_uses_first_or_latest():
    latest = HistoryRow(revision="P02", date="15.02.26", suitability="S2")
    result = DocumentResult(
        path=Path("sheet.pdf"),
        filename=FilenameFields(
            raw_stem="sheet",
            document_reference="ABC-WXY-ZZ-00-DR-A-0001",
            parse_ok=True,
        ),
        titleblock=TitleBlockFields(
            document_reference="ABC-WXY-ZZ-00-DR-A-0001",
            title="Roof Plan",
            revision="P02",
            date="01.03.26",
            history=RevisionHistory(
                rows=[
                    HistoryRow(revision="P01", date="01.01.26"),
                    latest,
                ],
                latest=latest,
            ),
        ),
        status=CheckStatus.HISTORY_MISMATCH,
        confidence=Confidence.REVIEW,
        issues=[CheckStatus.HISTORY_MISMATCH],
        history_comparisons=[
            FieldComparison("history_date", "01.03.26", "15.02.26", False, "neither"),
        ],
    )
    text = designer_actions(result)
    assert "01.03.26" in text
    assert "01.01.26" in text
    assert "15.02.26" in text
    assert "original issue" in text.lower()


def test_purpose_mismatch_p_construction():
    result = DocumentResult(
        path=Path("sheet.pdf"),
        filename=FilenameFields(raw_stem="sheet", revision="P01", parse_ok=True),
        titleblock=TitleBlockFields(
            document_reference="ABC-WXY-ZZ-00-DR-A-0001",
            title="Roof Plan",
            revision="P01",
            suitability="S5 - For Construction",
        ),
        status=CheckStatus.PURPOSE_MISMATCH,
        confidence=Confidence.REVIEW,
        issues=[CheckStatus.PURPOSE_MISMATCH],
    )
    text = designer_actions(result)
    assert "S5 - For Construction" in text
    assert "P01" in text
    assert "see bottom of this sheet for the approved list" in text.lower()
    assert "S3 Review and Comment" not in text


def test_purpose_issues_are_not_repeated():
    result = DocumentResult(
        path=Path("sheet.pdf"),
        filename=FilenameFields(
            raw_stem="sheet",
            revision="P01",
            parse_ok=True,
            parts={"project": "R459"},
        ),
        titleblock=TitleBlockFields(
            document_reference="ABC-WXY-ZZ-00-DR-A-0001",
            title="Roof Plan",
            revision="P01",
            suitability="S4 - Construction",
        ),
        status=CheckStatus.MULTIPLE_ISSUES,
        confidence=Confidence.REVIEW,
        issues=[
            CheckStatus.SUITABILITY_ERROR,
            CheckStatus.PURPOSE_MISMATCH,
        ],
        allowed_suitability=["S3 - For Review & Comment", "S5 - For Construction"],
    )
    text = designer_actions(result)
    assert not text.startswith("1. ")
    assert text.lower().count("purpose of issue is currently") == 1
    assert "see bottom of this sheet for the approved list" in text.lower()
    assert "mixes construction" not in text.lower()
    assert "S3 Review and Comment" not in text
    assert "Change it to" not in text
    assert "S5 - For Construction" not in text


def test_designer_sheet_lists_whitelist_not_a_predicted_status(tmp_path: Path):
    from drawing_qa.report import write_report

    result = DocumentResult(
        path=tmp_path / "R459-MBS-DZ-ZZ-DR-W-60001-P01.pdf",
        filename=FilenameFields(
            raw_stem="R459-MBS-DZ-ZZ-DR-W-60001-P01",
            document_reference="R459-MBS-DZ-ZZ-DR-W-60001",
            revision="P01",
            parse_ok=True,
            parts={"project": "R459"},
        ),
        titleblock=TitleBlockFields(
            document_reference="R459-MBS-DZ-ZZ-DR-W-60001",
            title="Roof Plan",
            revision="P01",
            suitability="S4 - Construction",
        ),
        status=CheckStatus.SUITABILITY_ERROR,
        confidence=Confidence.REVIEW,
        issues=[CheckStatus.SUITABILITY_ERROR],
        allowed_suitability=["S3 - For Review & Comment", "S5 - For Construction"],
        designer_purpose_values=["S3 - For Review & Comment", "S5 - For Construction"],
        purpose_list_official=True,
        purpose_list_name="Oval C+D",
    )
    output = write_report([result], tmp_path / "report.xlsx")
    from openpyxl import load_workbook

    from drawing_qa.paths import designer_report_path
    from drawing_qa.report import DESIGNER_HEADER_ROW

    sheet = load_workbook(output)["Designer actions"]
    values = [str(cell.value) for row in sheet.iter_rows(max_col=3) for cell in row if cell.value]
    assert sheet["A1"].value == "Designer actions — Oval C+D"
    assert sheet["B2"].value == "Oval C+D"
    assert sheet["A5"].value == "Need action"
    assert sheet.cell(DESIGNER_HEADER_ROW, 1).value == "Drawing number"
    assert "Approved purposes of issue" in values
    assert "Project R459 (Oval C+D)" in values
    assert "S3 - For Review & Comment" in values
    assert "S5 - For Construction" in values
    action = sheet.cell(DESIGNER_HEADER_ROW + 1, 3).value or ""
    assert "S4 - Construction" in action
    assert "see bottom of this sheet" in action.lower()
    assert "S3 Review and Comment" not in action
    side = load_workbook(designer_report_path(output))
    assert side.sheetnames == ["Designer actions"]
    assert side.active["A1"].value == sheet["A1"].value
    assert side.active.cell(DESIGNER_HEADER_ROW + 1, 3).value == sheet.cell(
        DESIGNER_HEADER_ROW + 1, 3
    ).value


def test_spelling_and_mismatch_are_numbered():
    result = build_result(
        _base(filename_rev="P01", titleblock_rev="P03", title="Groun Floor GA"),
        COMPARE_RULES,
    )
    result.spelling_errors = ["Groun"]
    result.issues.append(CheckStatus.SPELLING_ERROR)
    result.status = CheckStatus.MULTIPLE_ISSUES
    text = designer_actions(result)
    assert text.startswith("1. ")
    assert "2. " in text
    assert "Groun" in text
    assert "P03" in text


def test_incomplete_uses_filename_doc_ref_when_title_block_blank():
    result = DocumentResult(
        path=Path("sheet.pdf"),
        filename=FilenameFields(
            raw_stem="sheet",
            document_reference="ABC-WXY-ZZ-00-DR-A-0001",
            parse_ok=True,
        ),
        titleblock=TitleBlockFields(layout_id="bottom_right", title="Roof Plan"),
        status=CheckStatus.INCOMPLETE,
        confidence=Confidence.REVIEW,
        issues=[CheckStatus.INCOMPLETE],
    )
    assert designer_doc_ref(result) == "ABC-WXY-ZZ-00-DR-A-0001"
    assert "NUMBER" in designer_actions(result)


def _oval_list() -> list[str]:
    return ["S3 - For Review & Comment", "S4 - For Stage Approval", "S5 - For Construction"]


def test_history_note_is_not_mentioned_in_designer_actions():
    latest = HistoryRow(revision="C01", date="15.06.24", suitability="S3 - Bathroom first fix")
    result = DocumentResult(
        path=Path("sheet.pdf"),
        filename=FilenameFields(raw_stem="sheet", revision="C01", parse_ok=True),
        titleblock=TitleBlockFields(
            document_reference="ABC-WXY-ZZ-00-DR-A-0001",
            title="Roof Plan",
            revision="C01",
            suitability="S4 - Construction",
            history=RevisionHistory(rows=[latest], latest=latest),
        ),
        status=CheckStatus.SUITABILITY_ERROR,
        confidence=Confidence.REVIEW,
        issues=[CheckStatus.SUITABILITY_ERROR],
        allowed_suitability=_oval_list(),
        history_comparisons=[
            FieldComparison(
                "history_suitability",
                "S4 - Construction",
                "S3 - Bathroom first fix",
                True,
                "latest history row is a note, not a purpose of issue",
            ),
        ],
    )
    text = designer_actions(result)
    assert "not on the approved list" in text.lower()
    assert "must match" not in text.lower()
    assert "Bathroom" not in text
    assert "mixes construction" not in text.lower()


def test_history_whitelist_status_must_match_in_designer_actions():
    latest = HistoryRow(revision="C01", date="15.06.24", suitability="S5 - For Construction")
    result = DocumentResult(
        path=Path("sheet.pdf"),
        filename=FilenameFields(raw_stem="sheet", revision="C01", parse_ok=True),
        titleblock=TitleBlockFields(
            document_reference="ABC-WXY-ZZ-00-DR-A-0001",
            title="Roof Plan",
            revision="C01",
            suitability="S4 - Construction",
            history=RevisionHistory(rows=[latest], latest=latest),
        ),
        status=CheckStatus.MULTIPLE_ISSUES,
        confidence=Confidence.REVIEW,
        issues=[CheckStatus.SUITABILITY_ERROR, CheckStatus.HISTORY_MISMATCH],
        allowed_suitability=_oval_list(),
        history_comparisons=[
            FieldComparison(
                "history_suitability",
                "S4 - Construction",
                "S5 - For Construction",
                False,
                "mismatch",
            ),
        ],
    )
    text = designer_actions(result)
    assert "not on the approved list" in text.lower()
    assert "S5 - For Construction" in text
    assert "must match" in text.lower()
    assert "mixes construction" not in text.lower()
