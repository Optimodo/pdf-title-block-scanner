from pathlib import Path

from openpyxl import load_workbook

from drawing_qa.cli import main
from drawing_qa.models import CheckStatus, DocumentResult, FilenameFields, TitleBlockFields
from drawing_qa.paths import REPORT_NAME
from drawing_qa.rename import apply_renames
from drawing_qa.report import write_report
from tests.pdf_fixtures import write_bottom_right_pdf


def test_apply_renames_records_original_and_result(tmp_path: Path):
    source = tmp_path / "old-name.pdf"
    source.write_bytes(b"%PDF-1.4")
    dwg = tmp_path / "old-name.dwg"
    dwg.write_text("")
    result = DocumentResult(
        path=source,
        filename=FilenameFields(raw_stem="old-name", parse_ok=False),
        titleblock=TitleBlockFields(document_reference="ABC-XYZ-ZZ-00-DR-M-1234"),
        suggested_filename="ABC-XYZ-ZZ-00-DR-M-1234_Floor Plan_P01.pdf",
        original_filename="old-name.pdf",
        paired_dwg=dwg,
        dwg_files_present=True,
    )
    stats = apply_renames([result], verbose=False)
    assert stats.renamed == 1
    assert result.rename_result == "Renamed"
    assert result.original_filename == "old-name.pdf"
    assert result.path.name == "ABC-XYZ-ZZ-00-DR-M-1234_Floor Plan_P01.pdf"
    assert (tmp_path / "ABC-XYZ-ZZ-00-DR-M-1234_Floor Plan_P01.pdf").is_file()
    assert (tmp_path / "ABC-XYZ-ZZ-00-DR-M-1234_Floor Plan_P01.dwg").is_file()
    assert not source.exists()
    assert any("Renamed from old-name.pdf" in note for note in result.notes)


def test_apply_renames_skips_when_target_exists(tmp_path: Path):
    source = tmp_path / "old-name.pdf"
    source.write_bytes(b"%PDF-1.4")
    (tmp_path / "taken.pdf").write_bytes(b"%PDF-1.4")
    result = DocumentResult(
        path=source,
        filename=FilenameFields(raw_stem="old-name", parse_ok=False),
        titleblock=TitleBlockFields(document_reference="ABC-XYZ-ZZ-00-DR-M-1234"),
        suggested_filename="taken.pdf",
        original_filename="old-name.pdf",
    )
    stats = apply_renames([result], verbose=False)
    assert stats.skipped == 1
    assert result.rename_result == "Not renamed — taken.pdf already exists"
    assert source.is_file()


def test_report_keeps_scanned_name_after_rename(tmp_path: Path):
    result = DocumentResult(
        path=tmp_path / "new-name.pdf",
        filename=FilenameFields(
            raw_stem="old-name",
            document_reference="ABC-XYZ-ZZ-00-DR-M-0001",
            parse_ok=True,
        ),
        titleblock=TitleBlockFields(
            document_reference="ABC-XYZ-ZZ-00-DR-M-1234",
            title="Floor Plan",
            revision="P01",
        ),
        status=CheckStatus.MISMATCH,
        original_filename="old-name.pdf",
        suggested_filename="new-name.pdf",
        rename_result="Renamed",
        notes=["Renamed from old-name.pdf to new-name.pdf"],
    )
    output = tmp_path / "report.xlsx"
    write_report([result], output)
    wb = load_workbook(output)
    sheet = wb["All documents"]
    headers = [cell.value for cell in sheet[1]]
    assert "File (as scanned)" in headers
    assert "New filename" in headers
    assert "Rename result" in headers
    file_col = headers.index("File (as scanned)") + 1
    new_col = headers.index("New filename") + 1
    result_col = headers.index("Rename result") + 1
    notes_col = headers.index("Notes") + 1
    assert sheet.cell(2, file_col).value == "old-name.pdf"
    assert sheet.cell(2, new_col).value == "new-name.pdf"
    assert sheet.cell(2, result_col).value == "Renamed"
    assert "Renamed from old-name.pdf to new-name.pdf" in sheet.cell(2, notes_col).value
    summary = wb["Summary"]
    summary_text = " ".join(
        str(cell.value) for row in summary.iter_rows() for cell in row if cell.value
    )
    assert "Renamed" in summary_text
    assert "File (as scanned) is the name at the start of this run." in summary_text


def test_standardize_cli_renames_and_reports(tmp_path: Path, monkeypatch):
    write_bottom_right_pdf(
        tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf",
        document_reference="ABC-WXY-ZZ-00-DR-A-0001",
        title="Ground Floor GA",
        revision="P01",
    )
    monkeypatch.setattr("drawing_qa.cli.app_dir", lambda: tmp_path)
    code = main(["--standardize-names", "--no-pause"])
    assert code == 0
    new_name = "ABC-WXY-ZZ-00-DR-A-0001_Ground Floor GA_P01.pdf"
    assert (tmp_path / new_name).is_file()
    assert not (tmp_path / "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf").exists()
    report = tmp_path / REPORT_NAME
    assert report.is_file()
    wb = load_workbook(report)
    sheet = wb["All documents"]
    headers = [cell.value for cell in sheet[1]]
    file_col = headers.index("File (as scanned)") + 1
    new_col = headers.index("New filename") + 1
    result_col = headers.index("Rename result") + 1
    assert sheet.cell(2, file_col).value == "ABC-WXY-ZZ-00-DR-A-0001-P01.pdf"
    assert sheet.cell(2, new_col).value == new_name
    assert str(sheet.cell(2, result_col).value).startswith("Renamed")
